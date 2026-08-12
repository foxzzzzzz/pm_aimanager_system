from __future__ import annotations

import hashlib
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from openpyxl.workbook.workbook import Workbook

from project_manager_api.domain.models import (
    CanonicalProjectDraft,
    MilestoneDefinition,
    PlanDateState,
    PlanVersionDraft,
    PlanWindow,
    ProductSpecItem,
    ProjectIdentity,
    ProjectMemberDraft,
    RaciRole,
)
from project_manager_api.imports.errors import WorkbookValidationError
from project_manager_api.imports.manifest import Marker, TemplateManifest
from project_manager_api.imports.report import ImportCounts, ImportReport, ParseResult

DATE_PATTERN = r"\d{4}[/-]\d{1,2}[/-]\d{1,2}"
DATE_RANGE_PATTERN = re.compile(rf"^\s*({DATE_PATTERN})\s*(?:-|~|～|至)\s*({DATE_PATTERN})\s*$")
FORMULA_REFERENCE_PATTERN = re.compile(r"(?:(?:'([^']+)')|([^'!&+]+))!\$?([A-Z]{1,3})\$?(\d+)")


class LyraTemplateV1Parser:
    def __init__(self, manifest: TemplateManifest) -> None:
        self.manifest = manifest

    def has_template_identity(self, workbook: Workbook) -> bool:
        return all(
            self._marker_matches(workbook, marker) for marker in self.manifest.identity_markers
        )

    def parse_workbook(self, workbook: Workbook, source_path: Path) -> ParseResult:
        self._validate_structure(workbook)
        source_sha256 = hashlib.sha256(source_path.read_bytes()).hexdigest()
        document_version, project = self._parse_identity(workbook)
        product_specs = self._parse_product_specs(workbook)
        members = self._parse_members(workbook)
        plan_versions = self._parse_plan_versions(workbook)
        milestones = self._parse_milestones(workbook)
        self._validate_semantics(workbook, members, milestones)
        active_plan_name = plan_versions[-1].name
        draft = CanonicalProjectDraft(
            template_id=self.manifest.template_id,
            template_version=self.manifest.template_version,
            document_version=document_version,
            source_sha256=source_sha256,
            project=project,
            product_specs=product_specs,
            members=members,
            milestones=milestones,
            plan_versions=plan_versions,
            active_plan_name=active_plan_name,
        )
        warnings = self._build_warnings(draft)
        report = ImportReport(
            template=self.manifest.identifier,
            source_sha256=source_sha256,
            counts=ImportCounts(
                product_specs=len(product_specs),
                members=len(members),
                milestones=len(milestones),
                plan_versions=len(plan_versions),
            ),
            warnings=warnings,
        )
        return ParseResult(draft=draft, report=report)

    def _validate_structure(self, workbook: Workbook) -> None:
        for sheet_name in self.manifest.required_sheets:
            if sheet_name not in workbook.sheetnames:
                raise WorkbookValidationError(f"missing required sheet: {sheet_name}")
        for marker in self.manifest.required_markers:
            if not self._marker_matches(workbook, marker):
                raise WorkbookValidationError(
                    f"invalid required marker at {marker.sheet}!{marker.cell}"
                )

        progress = self.manifest.progress
        sheet = workbook[progress.sheet]
        for index, milestone in enumerate(self.manifest.milestones):
            column = progress.first_milestone_column + index
            actual = _cell_text(sheet.cell(progress.milestone_header_row, column).value)
            if actual != milestone.name:
                location = (
                    f"{progress.sheet}!{get_column_letter(column)}{progress.milestone_header_row}"
                )
                raise WorkbookValidationError(
                    f"invalid milestone header at {location}: expected {milestone.name!r}"
                )

    @staticmethod
    def _marker_matches(workbook: Workbook, marker: Marker) -> bool:
        if marker.sheet not in workbook.sheetnames:
            return False
        value = _cell_text(workbook[marker.sheet][marker.cell].value)
        if marker.equals is not None and value != marker.equals:
            return False
        return not (marker.contains is not None and marker.contains not in (value or ""))

    def _parse_identity(self, workbook: Workbook) -> tuple[str, ProjectIdentity]:
        sheet = workbook[self.manifest.product_specs.sheet]
        title = _cell_text(sheet["A1"].value) or ""
        product = _cell_text(sheet["D4"].value) or ""
        document_match = re.search(r"---\s*(V[\d.]+)", title, flags=re.IGNORECASE)
        code_match = re.search(r"项目号[：:]\s*([^；;)）]+)", product)
        name = re.split(r"[（(]", product, maxsplit=1)[0].strip()
        if document_match is None:
            raise WorkbookValidationError("missing document version at 产品规格书!A1")
        if code_match is None or not name:
            raise WorkbookValidationError("missing project identity at 产品规格书!D4")
        return document_match.group(1).upper(), ProjectIdentity(
            code=code_match.group(1).strip(),
            name=name,
        )

    def _parse_product_specs(self, workbook: Workbook) -> list[ProductSpecItem]:
        config = self.manifest.product_specs
        sheet = workbook[config.sheet]
        current_major: str | None = None
        current_category: str | None = None
        current_item: str | None = None
        items: list[ProductSpecItem] = []
        for row_number in range(config.first_row, config.last_row + 1):
            values = {
                key: _cell_text(sheet[f"{column}{row_number}"].value)
                for key, column in config.columns.items()
            }
            current_major = values["major_category"] or current_major
            current_category = values["category"] or current_category
            item_name = (
                values["item"] or values["category"] or values["major_category"] or current_item
            )
            if item_name is None:
                location = f"{config.sheet}!A{row_number}:C{row_number}"
                raise WorkbookValidationError(f"missing product specification item at {location}")
            current_item = item_name
            items.append(
                ProductSpecItem(
                    row_number=row_number,
                    major_category=current_major,
                    category=current_category,
                    item=item_name,
                    configuration=values["configuration"],
                    core_information=values["core_information"],
                    selected_model=values["selected_model"],
                    notes=values["notes"],
                    check_confirmation=values["check_confirmation"],
                    check_content=values["check_content"],
                )
            )
        return items

    def _parse_members(self, workbook: Workbook) -> list[ProjectMemberDraft]:
        config = self.manifest.team
        sheet = workbook[config.sheet]
        members: list[ProjectMemberDraft] = []
        member_indexes: dict[str, int] = {}
        for row_number in range(config.first_row, config.last_row + 1):
            values = {
                key: _cell_text(sheet[f"{column}{row_number}"].value)
                for key, column in config.columns.items()
            }
            if not values["role"] or not values["name"]:
                raise WorkbookValidationError(
                    f"missing team role or owner at {config.sheet}!B{row_number}:C{row_number}"
                )
            member = ProjectMemberDraft.model_validate(
                {
                    **values,
                    "is_project_manager": _has_role(
                        values["role"], self.manifest.team.project_manager_role
                    ),
                }
            )
            existing_index = member_indexes.get(member.name)
            if existing_index is None:
                member_indexes[member.name] = len(members)
                members.append(member)
                continue
            existing = members[existing_index]
            if _conflicting_value(existing.phone, member.phone) or _conflicting_value(
                existing.email, member.email
            ):
                raise WorkbookValidationError(
                    f"conflicting contact details for duplicate member at "
                    f"{config.sheet}!C{row_number}"
                )
            members[existing_index] = existing.model_copy(
                update={
                    "role": _merge_distinct(existing.role, member.role),
                    "is_project_manager": (
                        existing.is_project_manager or member.is_project_manager
                    ),
                    "phone": existing.phone or member.phone,
                    "email": existing.email or member.email,
                    "notes": _merge_distinct(existing.notes, member.notes),
                }
            )
        return members

    def _parse_plan_versions(self, workbook: Workbook) -> list[PlanVersionDraft]:
        progress = self.manifest.progress
        sheet = workbook[progress.sheet]
        versions: list[PlanVersionDraft] = []
        for row_number in range(progress.plan_first_row, progress.plan_last_row + 1):
            name = _cell_text(sheet.cell(row_number, 1).value)
            if not name:
                continue
            milestone_windows: dict[str, PlanWindow] = {}
            for index, milestone in enumerate(self.manifest.milestones):
                column = progress.first_milestone_column + index
                location = f"{progress.sheet}!{get_column_letter(column)}{row_number}"
                milestone_windows[milestone.name] = _parse_plan_window(
                    sheet.cell(row_number, column).value,
                    workbook.epoch,
                    location,
                )
            versions.append(PlanVersionDraft(name=name, milestones=milestone_windows))
        if not versions:
            raise WorkbookValidationError(f"no plan versions found in {progress.sheet}")
        return versions

    def _parse_milestones(self, workbook: Workbook) -> list[MilestoneDefinition]:
        progress = self.manifest.progress
        sheet = workbook[progress.sheet]
        milestones: list[MilestoneDefinition] = []
        for index, configured in enumerate(self.manifest.milestones):
            column = progress.first_milestone_column + index
            assignments = {
                RaciRole(role): self._resolve_assignment(
                    workbook,
                    sheet.cell(row_number, column).value,
                    f"{progress.sheet}!{get_column_letter(column)}{row_number}",
                )
                for role, row_number in progress.raci_rows.items()
            }
            variance_value = sheet.cell(progress.variance_days_row, column).value
            variance_days = int(variance_value) if isinstance(variance_value, int | float) else None
            milestones.append(
                MilestoneDefinition(
                    code=configured.code,
                    name=configured.name,
                    output=_cell_text(sheet.cell(progress.output_row, column).value),
                    actual_completion=_parse_plan_window(
                        sheet.cell(progress.actual_completion_row, column).value,
                        workbook.epoch,
                        f"{progress.sheet}!{get_column_letter(column)}{progress.actual_completion_row}",
                    ),
                    variance_days=variance_days,
                    variance_note=_cell_text(sheet.cell(progress.variance_note_row, column).value),
                    risk_note=_cell_text(sheet.cell(progress.risk_note_row, column).value),
                    assignments=assignments,
                )
            )
        return milestones

    @staticmethod
    def _resolve_assignment(workbook: Workbook, value: Any, location: str) -> list[str]:
        if value is None:
            return []
        if isinstance(value, str) and value.startswith("="):
            resolved: list[str] = []
            for match in FORMULA_REFERENCE_PATTERN.finditer(value):
                sheet_name = (match.group(1) or match.group(2)).lstrip("= ")
                cell_address = f"{match.group(3)}{match.group(4)}"
                if sheet_name not in workbook.sheetnames:
                    raise WorkbookValidationError(
                        f"invalid RACI formula reference at {location}: {sheet_name}!{cell_address}"
                    )
                member_name = _cell_text(workbook[sheet_name][cell_address].value)
                if member_name and member_name not in resolved:
                    resolved.append(member_name)
            if not resolved:
                raise WorkbookValidationError(f"unresolved RACI formula at {location}")
            return resolved
        text = _cell_text(value)
        return [part.strip() for part in re.split(r"[、,，;；]", text or "") if part.strip()]

    @staticmethod
    def _build_warnings(draft: CanonicalProjectDraft) -> list[str]:
        tbd_count = sum(
            window.state is PlanDateState.TBD for window in draft.active_plan.milestones.values()
        )
        return (
            [f"candidate current plan contains {tbd_count} TBD milestone(s)"] if tbd_count else []
        )

    def _validate_semantics(
        self,
        workbook: Workbook,
        members: list[ProjectMemberDraft],
        milestones: list[MilestoneDefinition],
    ) -> None:
        member_names = [member.name for member in members]
        milestone_codes = [milestone.code for milestone in milestones]
        if len(milestone_codes) != len(set(milestone_codes)):
            raise WorkbookValidationError("duplicate milestone codes in template manifest")
        managers = [member.name for member in members if member.is_project_manager]
        if len(managers) != 1:
            raise WorkbookValidationError(
                "exactly one project manager is required in "
                f"{self.manifest.team.sheet}; found {len(managers)}"
            )

        team_scope = _cell_text(workbook[self.manifest.team.sheet]["A4"].value)
        allowed_assignees = set(member_names)
        if team_scope:
            allowed_assignees.add(team_scope)
        progress = self.manifest.progress
        for index, milestone in enumerate(milestones):
            column = progress.first_milestone_column + index
            for role, assignees in milestone.assignments.items():
                unknown = [name for name in assignees if name not in allowed_assignees]
                if unknown:
                    row = progress.raci_rows[role.value]
                    location = f"{progress.sheet}!{get_column_letter(column)}{row}"
                    raise WorkbookValidationError(
                        f"unknown RACI assignee at {location}: {', '.join(unknown)}"
                    )


def _cell_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _conflicting_value(left: str | None, right: str | None) -> bool:
    return bool(left and right and left != right)


def _merge_distinct(left: str | None, right: str | None) -> str | None:
    values = [value for value in (left, right) if value]
    return " / ".join(dict.fromkeys(values)) or None


def _has_role(value: str | None, expected: str) -> bool:
    return expected in {part.strip() for part in (value or "").split("/")}


def _parse_plan_window(value: Any, epoch: datetime, location: str) -> PlanWindow:
    if value is None or (isinstance(value, str) and not value.strip()):
        return PlanWindow(state=PlanDateState.TBD)
    if isinstance(value, str) and value.strip().upper() == "N/A":
        return PlanWindow(state=PlanDateState.NOT_APPLICABLE)
    if isinstance(value, datetime):
        parsed = value.date()
        return PlanWindow(state=PlanDateState.SCHEDULED, start_date=parsed, end_date=parsed)
    if isinstance(value, date):
        return PlanWindow(state=PlanDateState.SCHEDULED, start_date=value, end_date=value)
    if isinstance(value, int | float):
        parsed = from_excel(value, epoch).date()
        return PlanWindow(state=PlanDateState.SCHEDULED, start_date=parsed, end_date=parsed)
    if isinstance(value, str):
        range_match = DATE_RANGE_PATTERN.match(value)
        if range_match:
            start = _parse_date_text(range_match.group(1), location)
            end = _parse_date_text(range_match.group(2), location)
            return PlanWindow(state=PlanDateState.SCHEDULED, start_date=start, end_date=end)
        try:
            parsed = _parse_date_text(value, location)
        except ValueError as exc:
            raise WorkbookValidationError(str(exc)) from exc
        return PlanWindow(state=PlanDateState.SCHEDULED, start_date=parsed, end_date=parsed)
    raise WorkbookValidationError(f"unsupported plan date at {location}: {value!r}")


def _parse_date_text(value: str, location: str) -> date:
    match = re.fullmatch(r"\s*(\d{4})[/-](\d{1,2})[/-](\d{1,2})\s*", value)
    if match is None:
        raise ValueError(f"invalid plan date at {location}: {value!r}")
    try:
        return date(*(int(part) for part in match.groups()))
    except ValueError as exc:
        raise ValueError(f"invalid plan date at {location}: {value!r}") from exc
