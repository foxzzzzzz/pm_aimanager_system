import shutil
import uuid
from collections.abc import Iterator
from datetime import date, timedelta
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from project_manager_api.api.app import create_app
from project_manager_api.api.schemas import IssueUpdate
from project_manager_api.db.base import Base
from project_manager_api.db.models import (
    ChangeProposal,
    Issue,
    IssueStatus,
    MilestoneRuntimeState,
)
from project_manager_api.services.errors import ConflictError
from project_manager_api.services.projects import ProjectService, _issue_risk, milestone_risk
from project_manager_api.settings import AppSettings

ROOT = Path(__file__).resolve().parents[3]
WORKBOOK = ROOT / "tests" / "fixtures" / "lyra-template-v1" / "lyra_v1_sanitized.xlsx"
MANIFEST = ROOT / "config" / "templates" / "lyra_project_spec-v1.0.yaml"
PM_HEADERS = {"Authorization": "Bearer test-admin-token"}


@pytest.fixture
def workflow() -> Iterator[tuple[TestClient, Path]]:
    with TemporaryDirectory(dir=ROOT / "tmp") as directory:
        workdir = Path(directory)
        settings = AppSettings(
            database_url=f"sqlite:///{workdir / 'phase2.sqlite'}",
            manifest_paths=[MANIFEST],
            import_storage_path=workdir / "imports",
            max_import_size_bytes=20 * 1024 * 1024,
            admin_api_token="test-admin-token",
            admin_actor_id="pm-001",
            phone_hmac_key="test-phone-key",
        )
        app = create_app(settings)
        Base.metadata.create_all(app.state.engine)
        with TestClient(app) as client:
            yield client, workdir
        app.state.engine.dispose()


def _headers(key: str | None = None, actor: str = "pm-001") -> dict[str, str]:
    token = "test-admin-token" if actor == "pm-001" else "invalid-admin-token"
    headers = {"Authorization": f"Bearer {token}"}
    if key is not None:
        headers["X-Idempotency-Key"] = key
    return headers


def _create_project(client: TestClient, key: str = "create-lyra") -> dict[str, Any]:
    response = client.post(
        "/api/v1/projects",
        headers=_headers(key),
        json={"code": "ZPD1322", "name": "Lyra Pro"},
    )
    assert response.status_code == 201, response.json()
    return response.json()


def _set_proposal_submitter(client: TestClient, proposal_id: str, actor_id: str) -> None:
    with client.app.state.session_factory() as session:
        proposal = session.scalar(
            select(ChangeProposal).where(ChangeProposal.id == uuid.UUID(proposal_id))
        )
        assert proposal is not None
        proposal.submitted_by_actor_id = actor_id
        session.commit()


def _upload(
    client: TestClient,
    project_id: str,
    workbook: Path = WORKBOOK,
    key: str = "import-lyra",
) -> dict[str, Any]:
    with workbook.open("rb") as source:
        response = client.post(
            f"/api/v1/projects/{project_id}/imports",
            headers=_headers(key),
            files={
                "file": (
                    workbook.name,
                    source,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert response.status_code == 201, response.json()
    return response.json()


def _publish(
    client: TestClient,
    import_id: str,
    expected_version: int,
    key: str,
) -> Any:
    return client.post(
        f"/api/v1/imports/{import_id}/publish",
        headers=_headers(key),
        json={"expected_project_version": expected_version},
    )


def _published_project(client: TestClient) -> tuple[str, dict[str, Any]]:
    project = _create_project(client)
    project_id = str(project["id"])
    imported = _upload(client, project_id)
    response = _publish(client, str(imported["id"]), 0, "publish-v1")
    assert response.status_code == 200
    assert response.json()["version_number"] == 1
    return project_id, response.json()


def _create_approved_issue(
    client: TestClient, project_id: str, payload: dict[str, Any], key: str
) -> dict[str, Any]:
    proposal = client.post(
        f"/api/v1/projects/{project_id}/issues",
        headers=_headers(key),
        json=payload,
    )
    assert proposal.status_code == 201, proposal.json()
    approved = client.post(
        f"/api/v1/issue-create-proposals/{proposal.json()['id']}/approve",
        headers=_headers(f"{key}-approve"),
    )
    assert approved.status_code == 200, approved.json()
    issue_id = approved.json()["issue_id"]
    issues = client.get(f"/api/v1/projects/{project_id}/issues", headers=PM_HEADERS).json()
    return next(issue for issue in issues if issue["id"] == issue_id)


def _copy_with_active_plan_date(workdir: Path, filename: str, value: date) -> Path:
    changed = workdir / filename
    shutil.copyfile(WORKBOOK, changed)
    workbook = load_workbook(changed)
    workbook["项目进度表"]["B5"] = value
    workbook.save(changed)
    workbook.close()
    return changed


def test_project_creation_is_idempotent_and_scoped_to_creator(
    workflow: tuple[TestClient, Path],
) -> None:
    client, _ = workflow

    first = _create_project(client)
    second = _create_project(client)

    assert first["id"] == second["id"]
    assert client.get("/api/v1/projects", headers=PM_HEADERS).json()[0]["code"] == "ZPD1322"
    forbidden = client.get(
        f"/api/v1/projects/{first['id']}/dashboard",
        headers=_headers(actor="outsider"),
    )
    assert forbidden.status_code == 403


def test_empty_project_can_be_edited_or_deleted_but_published_project_is_protected(
    workflow: tuple[TestClient, Path],
) -> None:
    client, _ = workflow
    project = _create_project(client)
    project_id = str(project["id"])

    updated = client.patch(
        f"/api/v1/projects/{project_id}",
        headers=_headers("edit-empty-project"),
        json={"code": "ZPD1322-R", "name": "Lyra Pro Revised"},
    )
    assert updated.status_code == 200, updated.json()
    assert updated.json()["code"] == "ZPD1322-R"
    assert updated.json()["name"] == "Lyra Pro Revised"

    deleted = client.delete(
        f"/api/v1/projects/{project_id}", headers=_headers("delete-empty-project")
    )
    assert deleted.status_code == 204
    assert client.get("/api/v1/projects", headers=PM_HEADERS).json() == []

    published_project = _create_project(client, key="create-published-project")
    published_project_id = str(published_project["id"])
    imported = _upload(client, published_project_id, key="import-published-project")
    assert _publish(client, str(imported["id"]), 0, "publish-protected-project").status_code == 200
    protected = client.patch(
        f"/api/v1/projects/{published_project_id}",
        headers=_headers("edit-published-project"),
        json={"code": "ZPD1322-R", "name": "Lyra Pro Revised"},
    )
    assert protected.status_code == 409
    assert protected.json()["detail"] == "published projects cannot be edited"
    protected = client.delete(
        f"/api/v1/projects/{published_project_id}",
        headers=_headers("delete-published-project"),
    )
    assert protected.status_code == 409
    assert protected.json()["detail"] == "published projects cannot be deleted"


def test_workbook_can_create_a_project_and_rejects_existing_project_code(
    workflow: tuple[TestClient, Path],
) -> None:
    client, _ = workflow

    with WORKBOOK.open("rb") as source:
        created = client.post(
            "/api/v1/imports",
            headers=_headers("create-project-from-workbook"),
            files={
                "file": (
                    WORKBOOK.name,
                    source,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert created.status_code == 201, created.json()
    assert created.json()["project"]["code"] == "ZPD1322"
    assert created.json()["project"]["name"] == "Lyra Pro"
    assert created.json()["import"]["status"] == "validated"
    assert created.json()["import"]["project_id"] == created.json()["project"]["id"]

    with WORKBOOK.open("rb") as source:
        conflict = client.post(
            "/api/v1/imports",
            headers=_headers("duplicate-workbook-project"),
            files={
                "file": (
                    WORKBOOK.name,
                    source,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert conflict.status_code == 409
    assert conflict.json()["detail"] == (
        "project code already exists; select the existing project to import"
    )


def test_existing_project_import_rejects_a_name_mismatch(
    workflow: tuple[TestClient, Path],
) -> None:
    client, _ = workflow
    project = client.post(
        "/api/v1/projects",
        headers=_headers("create-name-mismatch-project"),
        json={"code": "ZPD1322", "name": "Incorrect Name"},
    ).json()

    with WORKBOOK.open("rb") as source:
        response = client.post(
            f"/api/v1/projects/{project['id']}/imports",
            headers=_headers("reject-name-mismatch"),
            files={
                "file": (
                    WORKBOOK.name,
                    source,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert response.status_code == 409
    assert response.json()["detail"] == "workbook project name does not match the target project"


def test_idempotency_key_rejects_a_different_request_body(
    workflow: tuple[TestClient, Path],
) -> None:
    client, _ = workflow
    _create_project(client, key="same-key")

    response = client.post(
        "/api/v1/projects",
        headers=_headers("same-key"),
        json={"code": "OTHER", "name": "Different request"},
    )

    assert response.status_code == 409
    assert response.json()["detail"] == "idempotency key was already used with another request"


def test_admin_api_rejects_spoofed_actor_header_and_invalid_bearer(
    workflow: tuple[TestClient, Path],
) -> None:
    client, _ = workflow

    spoofed = client.get("/api/v1/projects", headers={"X-Actor-Id": "pm-001"})
    invalid = client.get(
        "/api/v1/projects", headers={"Authorization": "Bearer invalid-admin-token"}
    )

    assert spoofed.status_code == 401
    assert invalid.status_code == 403


def test_import_diff_publish_history_and_dashboard(workflow: tuple[TestClient, Path]) -> None:
    client, workdir = workflow
    project = _create_project(client)
    project_id = str(project["id"])

    imported = _upload(client, project_id)

    assert imported["status"] == "validated"
    assert imported["report"]["counts"] == {
        "product_specs": 70,
        "members": 22,
        "milestones": 24,
        "plan_versions": 3,
    }
    assert imported["diff_count"] > 0
    assert len(list((workdir / "imports").glob("*.xlsx"))) == 1

    published = _publish(client, str(imported["id"]), 0, "publish-v1")
    assert published.status_code == 200
    assert published.json()["version_number"] == 1

    dashboard = client.get(f"/api/v1/projects/{project_id}/dashboard", headers=PM_HEADERS)
    assert dashboard.status_code == 200
    assert dashboard.json()["counts"] == {
        "members": 22,
        "milestones": 24,
        "product_specs": 70,
        "issues_open": 0,
    }
    versions = client.get(f"/api/v1/projects/{project_id}/versions", headers=PM_HEADERS)
    assert [item["version_number"] for item in versions.json()] == [1]


def test_project_dashboard_exposes_server_classified_tasks_and_issue_summary(
    workflow: tuple[TestClient, Path],
) -> None:
    client, _ = workflow
    project_id, _ = _published_project(client)
    business_date = date.fromisoformat(
        client.get(f"/api/v1/projects/{project_id}/dashboard", headers=PM_HEADERS).json()[
            "business_date"
        ]
    )
    _create_approved_issue(
        client,
        project_id,
        {
            "description": "项目看板问题",
            "impact": "验证问题风险摘要",
            "owner_name": "成员10",
            "accountable_names": ["成员02"],
            "consulted_names": ["成员03"],
            "informed_names": ["成员04"],
            "severity": "high",
            "due_date": (business_date + timedelta(days=7)).isoformat(),
        },
        "dashboard-issue",
    )

    response = client.get(f"/api/v1/projects/{project_id}/dashboard", headers=PM_HEADERS)

    assert response.status_code == 200
    dashboard = response.json()
    assert dashboard["business_date"] == business_date.isoformat()
    assert dashboard["tasks"]
    assert all(
        task.get("plan", {}).get("state") != "not_applicable"
        for task in dashboard["tasks"]
    )
    assert set(dashboard["tasks"][0]) >= {
        "code",
        "name",
        "plan",
        "assignments",
        "risk",
    }
    assert set(dashboard["tasks"][0]["assignments"]) == {"R", "A", "C", "I"}
    assert dashboard["tasks"][0]["risk"] in {"todo", "upcoming", "overdue", "completed"}
    assert dashboard["issues"][0]["description"] == "项目看板问题"
    assert dashboard["issues"][0]["risk"] == "upcoming"


def test_project_review_exposes_specs_roles_raci_and_tbd_without_contacts(
    workflow: tuple[TestClient, Path],
) -> None:
    client, _ = workflow
    project_id, _ = _published_project(client)

    response = client.get(f"/api/v1/projects/{project_id}/review", headers=PM_HEADERS)

    assert response.status_code == 200
    review = response.json()
    assert review["current_version_number"] == 1
    assert len(review["product_specs"]) == 70
    assert review["product_specs"][0]["item"]
    assert "check_confirmation" in review["product_specs"][0]
    assert "check_content" in review["product_specs"][0]
    assert len(review["members"]) == 22
    assert review["members"][0]["role"] == "项目经理"
    assert "phone" not in review["members"][0]
    assert "email" not in review["members"][0]
    assert len(review["milestones"]) == 24
    assert review["tbd_count"] == 3
    tbd = next(item for item in review["milestones"] if item["schedule"]["state"] == "tbd")
    assert tbd["assignments"]["R"]
    assert tbd["assignments"]["A"]


def test_duplicate_file_does_not_create_duplicate_official_version(
    workflow: tuple[TestClient, Path],
) -> None:
    client, _ = workflow
    project_id, first_version = _published_project(client)

    duplicate = _upload(client, project_id, key="import-duplicate")
    assert duplicate["diff_count"] == 0
    response = _publish(client, str(duplicate["id"]), 1, "publish-duplicate")

    assert response.status_code == 200
    assert response.json()["id"] == first_version["id"]
    versions = client.get(f"/api/v1/projects/{project_id}/versions", headers=PM_HEADERS)
    assert len(versions.json()) == 1


def test_import_rejects_workbook_for_a_different_project(
    workflow: tuple[TestClient, Path],
) -> None:
    client, workdir = workflow
    project = client.post(
        "/api/v1/projects",
        headers=_headers("create-other-project"),
        json={"code": "OTHER001", "name": "Other project"},
    ).json()

    with WORKBOOK.open("rb") as source:
        response = client.post(
            f"/api/v1/projects/{project['id']}/imports",
            headers=_headers("import-wrong-project"),
            files={
                "file": (
                    WORKBOOK.name,
                    source,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 409
    assert "project code" in response.json()["detail"]
    assert list((workdir / "imports").glob("*.xlsx")) == []


def test_stale_import_conflicts_with_newer_change_on_same_field(
    workflow: tuple[TestClient, Path],
) -> None:
    client, workdir = workflow
    project_id, _ = _published_project(client)
    august_first = _copy_with_active_plan_date(workdir, "august-first.xlsx", date(2026, 8, 1))
    august_second = _copy_with_active_plan_date(workdir, "august-second.xlsx", date(2026, 8, 2))
    stale = _upload(client, project_id, august_first, "import-august-first")
    winner = _upload(client, project_id, august_second, "import-august-second")

    assert _publish(client, str(winner["id"]), 1, "publish-v2").status_code == 200
    conflict = _publish(client, str(stale["id"]), 1, "publish-stale")

    assert conflict.status_code == 409
    assert any("正式立项" in path for path in conflict.json()["detail"]["conflict_paths"])
    persisted = client.get(f"/api/v1/imports/{stale['id']}", headers=PM_HEADERS)
    assert persisted.json()["status"] == "conflict"


def test_resaving_business_identical_workbook_has_no_semantic_diff(
    workflow: tuple[TestClient, Path],
) -> None:
    client, workdir = workflow
    project_id, _ = _published_project(client)
    resaved = workdir / "resaved.xlsx"
    shutil.copyfile(WORKBOOK, resaved)
    workbook = load_workbook(resaved)
    workbook.save(resaved)
    workbook.close()

    imported = _upload(client, project_id, resaved, "import-resaved")

    assert imported["diff_count"] == 0


def test_historical_workbook_cannot_report_false_publish_success(
    workflow: tuple[TestClient, Path],
) -> None:
    client, workdir = workflow
    project_id, _ = _published_project(client)
    changed = _copy_with_active_plan_date(workdir, "history-v2.xlsx", date(2026, 8, 2))
    imported = _upload(client, project_id, changed, "history-v2-import")
    assert _publish(client, imported["id"], 1, "history-v2-publish").status_code == 200
    historical = _upload(client, project_id, key="historical-import")

    response = _publish(client, historical["id"], 2, "historical-publish")

    assert response.status_code == 409
    assert "historical version" in response.json()["detail"]
    dashboard = client.get(f"/api/v1/projects/{project_id}/dashboard", headers=PM_HEADERS)
    assert dashboard.json()["current_version_number"] == 2


def test_cancelled_import_does_not_change_current_version(
    workflow: tuple[TestClient, Path],
) -> None:
    client, workdir = workflow
    project_id, _ = _published_project(client)
    changed = _copy_with_active_plan_date(workdir, "cancelled.xlsx", date(2026, 8, 3))
    imported = _upload(client, project_id, changed, "import-cancelled")

    response = client.post(
        f"/api/v1/imports/{imported['id']}/cancel",
        headers=_headers("cancel-import"),
    )

    assert response.status_code == 200
    dashboard = client.get(f"/api/v1/projects/{project_id}/dashboard", headers=PM_HEADERS)
    assert dashboard.json()["current_version_number"] == 1


def test_publish_failure_rolls_back_without_partial_version(
    workflow: tuple[TestClient, Path],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    client, _ = workflow
    project = _create_project(client)
    project_id = str(project["id"])
    imported = _upload(client, project_id)
    original_audit = ProjectService._audit

    def fail_published_audit(
        self: ProjectService,
        target_project_id: uuid.UUID,
        action: str,
        entity_type: str,
        entity_id: str,
        before: dict[str, Any] | None = None,
        after: dict[str, Any] | None = None,
        reason: str | None = None,
    ) -> None:
        if action == "import.published":
            raise RuntimeError("simulated audit failure")
        original_audit(
            self,
            target_project_id,
            action,
            entity_type,
            entity_id,
            before,
            after,
            reason,
        )

    monkeypatch.setattr(ProjectService, "_audit", fail_published_audit)
    with pytest.raises(RuntimeError, match="simulated audit failure"):
        _publish(client, str(imported["id"]), 0, "publish-failure")

    dashboard = client.get(f"/api/v1/projects/{project_id}/dashboard", headers=PM_HEADERS)
    versions = client.get(f"/api/v1/projects/{project_id}/versions", headers=PM_HEADERS)
    assert dashboard.json()["current_version_number"] == 0
    assert versions.json() == []


def test_progress_proposal_approval_is_optimistic_and_audited(
    workflow: tuple[TestClient, Path],
) -> None:
    client, _ = workflow
    project_id, _ = _published_project(client)
    proposal = client.post(
        f"/api/v1/projects/{project_id}/milestones/M01/progress-proposals",
        headers=_headers("proposal-m01"),
        json={
            "base_version_number": 1,
            "start_date": "2026-08-04",
            "end_date": "2026-08-04",
            "reason": "供应链确认后调整",
        },
    )
    assert proposal.status_code == 201
    _set_proposal_submitter(
        client, proposal.json()["id"], "mobile:responsible-member"
    )

    first = client.post(
        f"/api/v1/change-proposals/{proposal.json()['id']}/approve",
        headers=_headers("approve-m01"),
        json={"expected_project_version": 1},
    )
    second = client.post(
        f"/api/v1/change-proposals/{proposal.json()['id']}/approve",
        headers=_headers("approve-m01-again"),
        json={"expected_project_version": 1},
    )

    assert first.status_code == 200
    assert first.json()["version_number"] == 1
    assert second.status_code == 409
    versions = client.get(f"/api/v1/projects/{project_id}/versions", headers=PM_HEADERS)
    assert [item["version_number"] for item in versions.json()] == [1]
    review = client.get(f"/api/v1/projects/{project_id}/review", headers=PM_HEADERS)
    m01 = next(item for item in review.json()["milestones"] if item["code"] == "M01")
    assert m01["schedule"]["end_date"] == "2026-08-04"
    audit = client.get(f"/api/v1/projects/{project_id}/audit-logs", headers=PM_HEADERS)
    assert "change_proposal.approved" in [item["action"] for item in audit.json()]


def test_proposal_cannot_be_rebased_by_supplying_the_new_current_version(
    workflow: tuple[TestClient, Path],
) -> None:
    client, workdir = workflow
    project_id, _ = _published_project(client)
    proposal = client.post(
        f"/api/v1/projects/{project_id}/milestones/M01/progress-proposals",
        headers=_headers("stale-proposal"),
        json={
            "base_version_number": 1,
            "start_date": "2026-08-04",
            "end_date": "2026-08-04",
            "reason": "old baseline",
        },
    ).json()
    _set_proposal_submitter(client, proposal["id"], "mobile:responsible-member")
    changed = _copy_with_active_plan_date(workdir, "winner.xlsx", date(2026, 8, 2))
    winner = _upload(client, project_id, changed, "winner-import")
    assert _publish(client, winner["id"], 1, "winner-publish").status_code == 200

    response = client.post(
        f"/api/v1/change-proposals/{proposal['id']}/approve",
        headers=_headers("stale-proposal-approve"),
        json={"expected_project_version": 2},
    )

    assert response.status_code == 409
    dashboard = client.get(f"/api/v1/projects/{project_id}/dashboard", headers=PM_HEADERS)
    assert dashboard.json()["current_version_number"] == 2


def test_formal_import_previews_and_resets_overlapping_runtime_update(
    workflow: tuple[TestClient, Path],
) -> None:
    client, workdir = workflow
    project_id, _ = _published_project(client)
    proposal = client.post(
        f"/api/v1/projects/{project_id}/milestones/M01/progress-proposals",
        headers=_headers("runtime-before-import"),
        json={
            "base_version_number": 1,
            "start_date": "2026-08-04",
            "end_date": "2026-08-04",
            "reason": "运行期调整",
        },
    ).json()
    approved = client.post(
        f"/api/v1/change-proposals/{proposal['id']}/approve",
        headers=_headers("approve-runtime-before-import"),
        json={"expected_project_version": 1},
    )
    assert approved.status_code == 200

    changed = _copy_with_active_plan_date(workdir, "formal-v2.xlsx", date(2026, 8, 2))
    imported = _upload(client, project_id, changed, "formal-v2-import")
    runtime_diff = [
        item for item in imported["diff"] if item["path"] == "runtime_state[M01].schedule"
    ]
    assert len(runtime_diff) == 1
    assert runtime_diff[0]["operation"] == "removed"
    published = _publish(client, imported["id"], 1, "formal-v2-publish")
    assert published.status_code == 200
    assert published.json()["version_number"] == 2

    review = client.get(f"/api/v1/projects/{project_id}/review", headers=PM_HEADERS).json()
    m01 = next(item for item in review["milestones"] if item["code"] == "M01")
    assert m01["schedule"]["end_date"] == "2026-08-02"
    editable = client.get(
        f"/api/v1/projects/{project_id}/editable-data", headers=PM_HEADERS
    ).json()
    active_plan = next(
        item for item in editable["plan_versions"] if item["name"] == editable["active_plan_name"]
    )
    m01_name = next(item["name"] for item in editable["milestones"] if item["code"] == "M01")
    assert active_plan["milestones"][m01_name]["end_date"] == "2026-08-02"
    with client.app.state.session_factory() as session:
        assert session.scalar(select(MilestoneRuntimeState)) is None


def test_pending_runtime_proposal_conflicts_after_same_target_is_approved(
    workflow: tuple[TestClient, Path],
) -> None:
    client, _ = workflow
    project_id, _ = _published_project(client)
    first = client.post(
        f"/api/v1/projects/{project_id}/milestones/M01/progress-proposals",
        headers=_headers("runtime-first"),
        json={
            "base_version_number": 1,
            "start_date": "2026-08-04",
            "end_date": "2026-08-04",
            "reason": "首次调整",
        },
    ).json()
    second = client.post(
        f"/api/v1/projects/{project_id}/milestones/M01/progress-proposals",
        headers=_headers("runtime-second"),
        json={
            "base_version_number": 1,
            "start_date": "2026-08-05",
            "end_date": "2026-08-05",
            "reason": "并发调整",
        },
    ).json()

    assert client.post(
        f"/api/v1/change-proposals/{first['id']}/approve",
        headers=_headers("runtime-first-approve"),
        json={"expected_project_version": 1},
    ).status_code == 200
    conflict = client.post(
        f"/api/v1/change-proposals/{second['id']}/approve",
        headers=_headers("runtime-second-approve"),
        json={"expected_project_version": 1},
    )

    assert conflict.status_code == 409
    versions = client.get(f"/api/v1/projects/{project_id}/versions", headers=PM_HEADERS)
    assert [item["version_number"] for item in versions.json()] == [1]


def test_runtime_revision_rejects_stale_proposal_after_value_returns_to_original(
    workflow: tuple[TestClient, Path],
) -> None:
    client, _ = workflow
    project_id, _ = _published_project(client)
    original = client.get(
        f"/api/v1/projects/{project_id}/review", headers=PM_HEADERS
    ).json()
    original_m01 = next(item for item in original["milestones"] if item["code"] == "M01")
    original_schedule = original_m01["schedule"]
    stale = client.post(
        f"/api/v1/projects/{project_id}/milestones/M01/progress-proposals",
        headers=_headers("runtime-aba-stale"),
        json={
            "base_version_number": 1,
            "start_date": "2026-08-06",
            "end_date": "2026-08-06",
            "reason": "稍后审批的旧提案",
        },
    ).json()
    first = client.post(
        f"/api/v1/projects/{project_id}/milestones/M01/progress-proposals",
        headers=_headers("runtime-aba-first"),
        json={
            "base_version_number": 1,
            "start_date": "2026-08-04",
            "end_date": "2026-08-04",
            "reason": "先调整到新值",
        },
    ).json()
    assert client.post(
        f"/api/v1/change-proposals/{first['id']}/approve",
        headers=_headers("runtime-aba-first-approve"),
        json={"expected_project_version": 1},
    ).status_code == 200
    restore = client.post(
        f"/api/v1/projects/{project_id}/milestones/M01/progress-proposals",
        headers=_headers("runtime-aba-restore"),
        json={
            "base_version_number": 1,
            "start_date": original_schedule["start_date"],
            "end_date": original_schedule["end_date"],
            "reason": "恢复原值",
        },
    ).json()
    assert client.post(
        f"/api/v1/change-proposals/{restore['id']}/approve",
        headers=_headers("runtime-aba-restore-approve"),
        json={"expected_project_version": 1},
    ).status_code == 200

    response = client.post(
        f"/api/v1/change-proposals/{stale['id']}/approve",
        headers=_headers("runtime-aba-stale-approve"),
        json={"expected_project_version": 1},
    )

    assert response.status_code == 409
    assert "runtime revision" in response.json()["detail"]


def test_import_publish_rejects_runtime_state_changed_after_preview(
    workflow: tuple[TestClient, Path],
) -> None:
    client, workdir = workflow
    project_id, _ = _published_project(client)
    first = client.post(
        f"/api/v1/projects/{project_id}/milestones/M01/progress-proposals",
        headers=_headers("runtime-before-preview"),
        json={
            "base_version_number": 1,
            "start_date": "2026-08-04",
            "end_date": "2026-08-04",
            "reason": "形成第一版运行状态",
        },
    ).json()
    assert client.post(
        f"/api/v1/change-proposals/{first['id']}/approve",
        headers=_headers("runtime-before-preview-approve"),
        json={"expected_project_version": 1},
    ).status_code == 200
    changed = _copy_with_active_plan_date(workdir, "runtime-preview.xlsx", date(2026, 8, 2))
    imported = _upload(client, project_id, changed, "runtime-preview-import")

    second = client.post(
        f"/api/v1/projects/{project_id}/milestones/M01/progress-proposals",
        headers=_headers("runtime-after-preview"),
        json={
            "base_version_number": 1,
            "start_date": "2026-08-05",
            "end_date": "2026-08-05",
            "reason": "预览后再次调整",
        },
    ).json()
    assert client.post(
        f"/api/v1/change-proposals/{second['id']}/approve",
        headers=_headers("runtime-after-preview-approve"),
        json={"expected_project_version": 1},
    ).status_code == 200

    response = _publish(client, imported["id"], 1, "runtime-stale-preview-publish")

    assert response.status_code == 409
    assert "runtime state changed" in response.json()["detail"]
    review = client.get(f"/api/v1/projects/{project_id}/review", headers=PM_HEADERS).json()
    assert review["current_version_number"] == 1
    m01 = next(item for item in review["milestones"] if item["code"] == "M01")
    assert m01["schedule"]["end_date"] == "2026-08-05"


def test_admin_baseline_change_previews_and_resets_runtime_schedule(
    workflow: tuple[TestClient, Path],
) -> None:
    client, _ = workflow
    project_id, _ = _published_project(client)
    proposal = client.post(
        f"/api/v1/projects/{project_id}/milestones/M01/progress-proposals",
        headers=_headers("runtime-before-admin-change"),
        json={
            "base_version_number": 1,
            "start_date": "2026-08-04",
            "end_date": "2026-08-04",
            "reason": "运行期调整",
        },
    ).json()
    assert client.post(
        f"/api/v1/change-proposals/{proposal['id']}/approve",
        headers=_headers("runtime-before-admin-change-approve"),
        json={"expected_project_version": 1},
    ).status_code == 200
    editable = client.get(
        f"/api/v1/projects/{project_id}/editable-data", headers=PM_HEADERS
    ).json()
    active_plan = next(
        item for item in editable["plan_versions"] if item["name"] == editable["active_plan_name"]
    )
    m01_name = next(item["name"] for item in editable["milestones"] if item["code"] == "M01")
    replacement = {**active_plan, "milestones": {**active_plan["milestones"]}}
    replacement["milestones"][m01_name] = {
        "state": "scheduled",
        "start_date": "2026-08-03",
        "end_date": "2026-08-03",
    }
    change_set = client.post(
        f"/api/v1/projects/{project_id}/change-sets",
        headers=_headers("admin-runtime-reset-change-set"),
        json={
            "base_version_number": 1,
            "reason": "正式基线调整",
            "operations": [
                {
                    "op": "replace",
                    "resource": "plan",
                    "key": active_plan["name"],
                    "value": replacement,
                }
            ],
        },
    )
    assert change_set.status_code == 201
    assert any(
        item["path"] == "runtime_state[M01].schedule"
        for item in change_set.json()["diff"]
    )

    published = client.post(
        f"/api/v1/change-sets/{change_set.json()['id']}/publish",
        headers=_headers("admin-runtime-reset-publish"),
        json={"expected_project_version": 1},
    )

    assert published.status_code == 200
    assert published.json()["version_number"] == 2
    review = client.get(f"/api/v1/projects/{project_id}/review", headers=PM_HEADERS).json()
    m01 = next(item for item in review["milestones"] if item["code"] == "M01")
    assert m01["schedule"]["end_date"] == "2026-08-03"


def test_schedule_proposal_rejects_reversed_date_range(
    workflow: tuple[TestClient, Path],
) -> None:
    client, _ = workflow
    project_id, _ = _published_project(client)

    response = client.post(
        f"/api/v1/projects/{project_id}/milestones/M01/progress-proposals",
        headers=_headers("invalid-range"),
        json={
            "base_version_number": 1,
            "start_date": "2026-09-10",
            "end_date": "2026-09-01",
            "reason": "invalid range",
        },
    )

    assert response.status_code == 422


def test_issue_updates_require_current_revision_and_create_audit(
    workflow: tuple[TestClient, Path],
) -> None:
    client, _ = workflow
    project_id, _ = _published_project(client)
    created = _create_approved_issue(
        client,
        project_id,
        {
            "description": "摄像头调试存在低照噪点",
            "impact": "影响DVT验收",
            "owner_name": "成员10",
            "accountable_names": ["成员02"],
            "severity": "high",
            "due_date": "2026-08-20",
        },
        "issue-create",
    )

    updated = client.patch(
        f"/api/v1/issues/{created['id']}",
        headers=_headers("issue-update"),
        json={"expected_revision": 1, "status": "处理中"},
    )
    stale = client.patch(
        f"/api/v1/issues/{created['id']}",
        headers=_headers("issue-update-stale"),
        json={"expected_revision": 1, "status": "已解决"},
    )

    assert updated.status_code == 200
    assert updated.json()["revision"] == 2
    assert stale.status_code == 409
    issues = client.get(f"/api/v1/projects/{project_id}/issues", headers=PM_HEADERS)
    assert issues.json()[0]["status"] == "处理中"
    audit = client.get(f"/api/v1/projects/{project_id}/audit-logs", headers=PM_HEADERS)
    assert "issue.updated" in [item["action"] for item in audit.json()]

    requested = client.request(
        "DELETE",
        f"/api/v1/issues/{created['id']}",
        headers=_headers("issue-delete"),
        json={"expected_revision": 2, "reason": "问题记录作废"},
    )
    assert requested.status_code == 201
    assert requested.json()["status"] == "pending"
    duplicate = client.request(
        "DELETE",
        f"/api/v1/issues/{created['id']}",
        headers=_headers("issue-delete-duplicate"),
        json={"expected_revision": 2, "reason": "重复申请"},
    )
    assert duplicate.status_code == 409
    unchanged = client.get(f"/api/v1/projects/{project_id}/issues", headers=PM_HEADERS)
    assert unchanged.json()[0]["revision"] == 2
    assert unchanged.json()[0]["status"] != IssueStatus.CLOSED
    dashboard = client.get(f"/api/v1/projects/{project_id}/dashboard", headers=PM_HEADERS)
    assert dashboard.json()["counts"]["issues_open"] == 1

    queue = client.get(
        f"/api/v1/projects/{project_id}/issue-delete-proposals", headers=PM_HEADERS
    )
    assert [item["id"] for item in queue.json()] == [requested.json()["id"]]
    approved = client.post(
        f"/api/v1/issue-delete-proposals/{requested.json()['id']}/approve",
        headers=_headers("issue-delete-approve"),
    )
    assert approved.status_code == 200
    assert approved.json()["status"] == "approved"
    deleted = client.get(f"/api/v1/projects/{project_id}/issues", headers=PM_HEADERS)
    assert deleted.json()[0]["status"] == IssueStatus.CLOSED
    assert deleted.json()[0]["revision"] == 3
    dashboard = client.get(f"/api/v1/projects/{project_id}/dashboard", headers=PM_HEADERS)
    assert dashboard.json()["counts"]["issues_open"] == 0
    audit = client.get(f"/api/v1/projects/{project_id}/audit-logs", headers=PM_HEADERS)
    assert "issue.deleted" in [item["action"] for item in audit.json()]
    assert "issue_delete_proposal.created" in [item["action"] for item in audit.json()]
    assert "issue_delete_proposal.approved" in [item["action"] for item in audit.json()]


def test_rejected_issue_delete_proposal_keeps_issue_active(
    workflow: tuple[TestClient, Path],
) -> None:
    client, _ = workflow
    project_id, _ = _published_project(client)
    issue = _create_approved_issue(
        client,
        project_id,
        {
            "description": "Delete approval rejection",
            "impact": "Issue must remain active",
            "owner_name": "成员10",
            "accountable_names": ["成员02"],
            "severity": "high",
            "due_date": "2026-08-20",
        },
        "issue-delete-reject-create",
    )
    requested = client.request(
        "DELETE",
        f"/api/v1/issues/{issue['id']}",
        headers=_headers("issue-delete-reject-request"),
        json={"expected_revision": 1, "reason": "mistaken entry"},
    )
    rejected = client.post(
        f"/api/v1/issue-delete-proposals/{requested.json()['id']}/reject",
        headers=_headers("issue-delete-reject"),
        json={"reason": "keep tracking"},
    )

    assert rejected.status_code == 200
    assert rejected.json()["status"] == "rejected"
    remaining = client.get(f"/api/v1/projects/{project_id}/issues", headers=PM_HEADERS)
    assert remaining.json()[0]["revision"] == 1
    assert remaining.json()[0]["status"] != IssueStatus.CLOSED


def test_issue_raci_requires_project_members_and_returns_derived_risk(
    workflow: tuple[TestClient, Path],
) -> None:
    client, _ = workflow
    project_id, _ = _published_project(client)
    invalid = client.post(
        f"/api/v1/projects/{project_id}/issues",
        headers=_headers("issue-raci-invalid"),
        json={
            "description": "RACI成员校验",
            "impact": "避免错误通知",
            "owner_name": "成员10",
            "accountable_names": ["不存在成员"],
            "consulted_names": ["成员03"],
            "informed_names": [],
            "severity": "high",
            "due_date": "2026-08-20",
        },
    )
    assert invalid.status_code == 409

    created = _create_approved_issue(
        client,
        project_id,
        {
            "description": "RACI完整问题",
            "impact": "验证协作角色",
            "owner_name": "成员10",
            "accountable_names": ["成员02"],
            "consulted_names": ["成员03", "成员04"],
            "informed_names": ["成员05"],
            "severity": "high",
            "due_date": "2026-08-20",
        },
        "issue-raci-valid",
    )
    assert created["accountable_names"] == ["成员02"]
    assert created["consulted_names"] == ["成员03", "成员04"]
    assert created["informed_names"] == ["成员05"]
    assert created["risk"] in {"todo", "upcoming", "overdue"}


def test_issue_risk_uses_inclusive_fourteen_day_boundary() -> None:
    business_date = date(2026, 8, 12)
    issue = Issue(due_date=business_date, status=IssueStatus.OPEN)

    assert _issue_risk(issue, business_date, 14) == "upcoming"
    issue.due_date = date(2026, 8, 26)
    assert _issue_risk(issue, business_date, 14) == "upcoming"
    issue.due_date = date(2026, 8, 27)
    assert _issue_risk(issue, business_date, 14) == "todo"
    issue.due_date = date(2026, 8, 11)
    assert _issue_risk(issue, business_date, 14) == "overdue"
    issue.status = IssueStatus.RESOLVED
    assert _issue_risk(issue, business_date, 14) == "completed"


def test_milestone_risk_uses_the_same_inclusive_fourteen_day_boundary() -> None:
    business_date = date(2026, 8, 12)
    milestone = {
        "actual_completion": {"end_date": None},
        "plan": {"state": "scheduled", "end_date": "2026-08-26"},
    }

    assert milestone_risk(milestone, business_date, 14) == "upcoming"
    milestone["plan"]["end_date"] = "2026-08-27"
    assert milestone_risk(milestone, business_date, 14) == "todo"
    milestone["plan"]["end_date"] = "2026-08-11"
    assert milestone_risk(milestone, business_date, 14) == "overdue"
    milestone["actual_completion"]["end_date"] = "2026-08-10"
    assert milestone_risk(milestone, business_date, 14) == "completed"


def test_issue_revision_check_uses_the_current_database_value(
    workflow: tuple[TestClient, Path],
) -> None:
    client, _ = workflow
    project_id, _ = _published_project(client)
    created = _create_approved_issue(
        client,
        project_id,
        {
            "description": "并发更新验证",
            "impact": "验证乐观锁",
            "owner_name": "成员10",
            "accountable_names": ["成员02"],
            "severity": "high",
            "due_date": "2026-08-20",
        },
        "issue-atomic-create",
    )
    issue_id = uuid.UUID(created["id"])
    stale_session = client.app.state.session_factory()
    try:
        stale_issue = stale_session.get(Issue, issue_id)
        assert stale_issue is not None and stale_issue.revision == 1
        updated = client.patch(
            f"/api/v1/issues/{issue_id}",
            headers=_headers("issue-atomic-first"),
            json={"expected_revision": 1, "status": "处理中"},
        )
        assert updated.status_code == 200

        with pytest.raises(ConflictError, match="revision is stale"):
            ProjectService(stale_session, "pm-001").update_issue_as_member(
                stale_issue,
                IssueUpdate(expected_revision=1, status="已解决"),
            )
    finally:
        stale_session.rollback()
        stale_session.close()


def test_admin_change_set_crud_publishes_one_immutable_version(
    workflow: tuple[TestClient, Path],
) -> None:
    client, _ = workflow
    project_id, _ = _published_project(client)
    editable = client.get(
        f"/api/v1/projects/{project_id}/editable-data", headers=PM_HEADERS
    )
    assert editable.status_code == 200
    assert editable.json()["current_version_number"] == 1
    original_specs = editable.json()["product_specs"]
    first = original_specs[0]
    removed = original_specs[1]

    created = client.post(
        f"/api/v1/projects/{project_id}/change-sets",
        headers=_headers("change-set-spec-crud"),
        json={
            "base_version_number": 1,
            "reason": "项目核对修正产品规格",
            "operations": [
                {
                    "op": "replace",
                    "resource": "product_spec",
                    "key": str(first["row_number"]),
                    "value": {**first, "configuration": "Web修正值"},
                },
                {
                    "op": "remove",
                    "resource": "product_spec",
                    "key": str(removed["row_number"]),
                },
                {
                    "op": "add",
                    "resource": "product_spec",
                    "key": "999",
                    "value": {
                        "row_number": 999,
                        "major_category": "新增大类",
                        "category": "新增分类",
                        "item": "Web新增规格",
                        "configuration": "新增值",
                        "core_information": None,
                        "selected_model": None,
                        "notes": None,
                        "check_confirmation": None,
                        "check_content": None,
                    },
                },
            ],
        },
    )

    assert created.status_code == 201, created.json()
    assert created.json()["status"] == "pending"
    assert created.json()["diff"]
    published = client.post(
        f"/api/v1/change-sets/{created.json()['id']}/publish",
        headers=_headers("publish-spec-crud"),
        json={"expected_project_version": 1},
    )
    assert published.status_code == 200
    assert published.json()["version_number"] == 2

    current = client.get(
        f"/api/v1/projects/{project_id}/editable-data", headers=PM_HEADERS
    ).json()
    by_row = {item["row_number"]: item for item in current["product_specs"]}
    assert by_row[first["row_number"]]["configuration"] == "Web修正值"
    assert removed["row_number"] not in by_row
    assert by_row[999]["item"] == "Web新增规格"
    versions = client.get(f"/api/v1/projects/{project_id}/versions", headers=PM_HEADERS)
    assert [item["version_number"] for item in versions.json()] == [1, 2]


def test_change_set_rejects_dangling_member_reference_and_stale_baseline(
    workflow: tuple[TestClient, Path],
) -> None:
    client, _ = workflow
    project_id, _ = _published_project(client)
    editable = client.get(
        f"/api/v1/projects/{project_id}/editable-data", headers=PM_HEADERS
    ).json()
    referenced_name = next(
        name
        for milestone in editable["milestones"]
        for names in milestone["assignments"].values()
        for name in names
    )

    invalid = client.post(
        f"/api/v1/projects/{project_id}/change-sets",
        headers=_headers("remove-referenced-member"),
        json={
            "base_version_number": 1,
            "reason": "验证成员引用保护",
            "operations": [
                {"op": "remove", "resource": "member", "key": referenced_name}
            ],
        },
    )
    assert invalid.status_code == 409
    assert "RACI" in invalid.json()["detail"]

    winner_response = client.post(
        f"/api/v1/projects/{project_id}/change-sets",
        headers=_headers("winner-change-set"),
        json={
            "base_version_number": 1,
            "reason": "创建新版本",
            "operations": [
                {
                    "op": "replace",
                    "resource": "product_spec",
                    "key": str(editable["product_specs"][0]["row_number"]),
                    "value": {
                        **editable["product_specs"][0],
                        "notes": "winner",
                    },
                }
            ],
        },
    )
    assert winner_response.status_code == 201, winner_response.json()
    winner = winner_response.json()
    assert client.post(
        f"/api/v1/change-sets/{winner['id']}/publish",
        headers=_headers("publish-winner-change-set"),
        json={"expected_project_version": 1},
    ).status_code == 200

    stale = client.post(
        f"/api/v1/projects/{project_id}/change-sets",
        headers=_headers("stale-change-set"),
        json={
            "base_version_number": 1,
            "reason": "过期基线",
            "operations": [
                {
                    "op": "replace",
                    "resource": "product_spec",
                    "key": str(editable["product_specs"][0]["row_number"]),
                    "value": editable["product_specs"][0],
                }
            ],
        },
    )
    assert stale.status_code == 409


def test_change_set_rejects_unknown_resource_fields(
    workflow: tuple[TestClient, Path],
) -> None:
    client, _ = workflow
    project_id, _ = _published_project(client)
    editable = client.get(
        f"/api/v1/projects/{project_id}/editable-data", headers=PM_HEADERS
    ).json()
    spec = editable["product_specs"][0]

    response = client.post(
        f"/api/v1/projects/{project_id}/change-sets",
        headers=_headers("unknown-change-set-field"),
        json={
            "base_version_number": 1,
            "reason": "验证字段白名单",
            "operations": [
                {
                    "op": "replace",
                    "resource": "product_spec",
                    "key": str(spec["row_number"]),
                    "value": {**spec, "unexpected": "must not persist"},
                }
            ],
        },
    )

    assert response.status_code == 409
    assert "unknown fields" in response.json()["detail"]


def test_cancelled_change_set_does_not_publish(workflow: tuple[TestClient, Path]) -> None:
    client, _ = workflow
    project_id, _ = _published_project(client)
    editable = client.get(
        f"/api/v1/projects/{project_id}/editable-data", headers=PM_HEADERS
    ).json()
    spec = editable["product_specs"][0]
    created_response = client.post(
        f"/api/v1/projects/{project_id}/change-sets",
        headers=_headers("cancelled-change-set"),
        json={
            "base_version_number": 1,
            "reason": "取消不应发布",
            "operations": [
                {
                    "op": "replace",
                    "resource": "product_spec",
                    "key": str(spec["row_number"]),
                    "value": {**spec, "notes": "should-not-publish"},
                }
            ],
        },
    )
    assert created_response.status_code == 201, created_response.json()
    created = created_response.json()

    cancelled = client.post(
        f"/api/v1/change-sets/{created['id']}/cancel",
        headers=_headers("cancel-change-set"),
    )

    assert cancelled.status_code == 200
    assert cancelled.json()["status"] == "cancelled"
    dashboard = client.get(f"/api/v1/projects/{project_id}/dashboard", headers=PM_HEADERS)
    assert dashboard.json()["current_version_number"] == 1


def test_change_set_atomically_removes_member_and_all_raci_references(
    workflow: tuple[TestClient, Path],
) -> None:
    client, _ = workflow
    project_id, _ = _published_project(client)
    editable = client.get(
        f"/api/v1/projects/{project_id}/editable-data", headers=PM_HEADERS
    ).json()
    referenced_name = next(
        name
        for milestone in editable["milestones"]
        for names in milestone["assignments"].values()
        for name in names
        if name
        in {
            member["name"]
            for member in editable["members"]
            if not member.get("is_project_manager")
        }
    )
    operations = [
        {
            "op": "replace",
            "resource": "raci",
            "key": milestone["code"],
            "value": {
                role: [name for name in names if name != referenced_name]
                for role, names in milestone["assignments"].items()
            },
        }
        for milestone in editable["milestones"]
        if any(referenced_name in names for names in milestone["assignments"].values())
    ]
    operations.append(
        {"op": "remove", "resource": "member", "key": referenced_name}
    )

    created = client.post(
        f"/api/v1/projects/{project_id}/change-sets",
        headers=_headers("remove-member-and-raci"),
        json={
            "base_version_number": 1,
            "reason": "成员离项并移除RACI引用",
            "operations": operations,
        },
    )
    assert created.status_code == 201, created.json()
    published = client.post(
        f"/api/v1/change-sets/{created.json()['id']}/publish",
        headers=_headers("publish-remove-member-and-raci"),
        json={"expected_project_version": 1},
    )
    assert published.status_code == 200

    current = client.get(
        f"/api/v1/projects/{project_id}/editable-data", headers=PM_HEADERS
    ).json()
    assert referenced_name not in {member["name"] for member in current["members"]}
    assert all(
        referenced_name not in names
        for milestone in current["milestones"]
        for names in milestone["assignments"].values()
    )


def test_change_set_adds_milestone_to_every_plan_version(
    workflow: tuple[TestClient, Path],
) -> None:
    client, _ = workflow
    project_id, _ = _published_project(client)
    editable = client.get(
        f"/api/v1/projects/{project_id}/editable-data", headers=PM_HEADERS
    ).json()
    window = {"state": "tbd", "start_date": None, "end_date": None}
    operations: list[dict[str, Any]] = [
        {
            "op": "add",
            "resource": "milestone",
            "key": "M99",
            "value": {
                "code": "M99",
                "name": "Web验收节点",
                "output": "验收记录",
                "actual_completion": window,
                "variance_days": None,
                "variance_note": None,
                "risk_note": None,
                "assignments": {"R": [], "A": [], "C": [], "I": []},
            },
        }
    ]
    operations.extend(
        {
            "op": "replace",
            "resource": "plan",
            "key": plan["name"],
            "value": {
                **plan,
                "milestones": {**plan["milestones"], "Web验收节点": window},
            },
        }
        for plan in editable["plan_versions"]
    )
    created = client.post(
        f"/api/v1/projects/{project_id}/change-sets",
        headers=_headers("add-milestone-all-plans"),
        json={
            "base_version_number": 1,
            "reason": "新增验收节点",
            "operations": operations,
        },
    )
    assert created.status_code == 201, created.json()
    assert client.post(
        f"/api/v1/change-sets/{created.json()['id']}/publish",
        headers=_headers("publish-add-milestone"),
        json={"expected_project_version": 1},
    ).status_code == 200
    current = client.get(
        f"/api/v1/projects/{project_id}/editable-data", headers=PM_HEADERS
    ).json()
    assert "M99" in {item["code"] for item in current["milestones"]}
    assert all(
        "Web验收节点" in plan["milestones"] for plan in current["plan_versions"]
    )
