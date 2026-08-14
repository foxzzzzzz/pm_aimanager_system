from __future__ import annotations

import shutil
import uuid
from collections.abc import Iterator
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from project_manager_api.api.app import create_app
from project_manager_api.db.base import Base
from project_manager_api.db.models import (
    IdempotencyRecord,
    MemberBinding,
    MobileUser,
    ProjectMembership,
    ProjectRole,
    ProjectVersion,
)
from project_manager_api.settings import AppSettings

ROOT = Path(__file__).resolve().parents[3]
WORKBOOK = ROOT / "tests" / "fixtures" / "lyra-template-v1" / "lyra_v1_sanitized.xlsx"
MANIFEST = ROOT / "config" / "templates" / "lyra_project_spec-v1.0.yaml"
PM = {"Authorization": "Bearer test-admin-token"}


@pytest.fixture
def mobile_workflow() -> Iterator[TestClient]:
    with TemporaryDirectory(dir=ROOT / "tmp") as directory:
        workdir = Path(directory)
        settings = AppSettings(
            database_url=f"sqlite:///{workdir / 'phase3.sqlite'}",
            manifest_paths=[MANIFEST],
            import_storage_path=workdir / "imports",
            max_import_size_bytes=20 * 1024 * 1024,
            allow_development_wechat_login=True,
            admin_api_token="test-admin-token",
            admin_actor_id="pm-001",
            phone_hmac_key="test-phone-key",
            phone_encryption_key="AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA",
            wechat_subscription_template_id="test-template",
        )
        app = create_app(settings)
        Base.metadata.create_all(app.state.engine)
        with TestClient(app) as client:
            yield client
        app.state.engine.dispose()


def _admin_headers(key: str) -> dict[str, str]:
    return {**PM, "X-Idempotency-Key": key}


def _published_project(client: TestClient) -> str:
    project = client.post(
        "/api/v1/projects",
        headers=_admin_headers("phase3-create-project"),
        json={"code": "ZPD1322", "name": "Lyra Pro"},
    ).json()
    with WORKBOOK.open("rb") as source:
        imported = client.post(
            f"/api/v1/projects/{project['id']}/imports",
            headers=_admin_headers("phase3-import"),
            files={
                "file": (
                    WORKBOOK.name,
                    source,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        ).json()
    published = client.post(
        f"/api/v1/imports/{imported['id']}/publish",
        headers=_admin_headers("phase3-publish"),
        json={"expected_project_version": 0},
    )
    assert published.status_code == 200
    return str(project["id"])


def _login(client: TestClient, code: str) -> tuple[dict[str, str], dict[str, Any]]:
    response = client.post(
        "/api/v1/mobile/auth/wechat",
        json={"code": code, "display_name": code},
    )
    assert response.status_code == 200
    payload = response.json()
    return {"Authorization": f"Bearer {payload['access_token']}"}, payload


def _invite(
    client: TestClient,
    project_id: str,
    member_name: str,
    key: str,
    expected_phone: str | None = None,
) -> dict[str, Any]:
    response = client.post(
        f"/api/v1/projects/{project_id}/member-invitations",
        headers=_admin_headers(key),
        json={"member_name": member_name, "expected_phone": expected_phone},
    )
    assert response.status_code == 201
    invitation = response.json()
    assert len(invitation["invitation_token"]) <= 32
    assert invitation["mini_program_path"] == (
        f"pages/index/index?invitation={invitation['invitation_token']}"
    )
    return invitation


def test_unbound_user_cannot_view_project_and_bound_user_can(
    mobile_workflow: TestClient,
) -> None:
    client = mobile_workflow
    project_id = _published_project(client)
    outsider_headers, _ = _login(client, "dev:outsider")

    assert client.get("/api/v1/mobile/projects", headers=outsider_headers).json() == []
    forbidden = client.get(
        f"/api/v1/mobile/projects/{project_id}/dashboard", headers=outsider_headers
    )
    assert forbidden.status_code == 403

    invitation = _invite(client, project_id, "成员10", "invite-member10")
    with client.app.state.session_factory() as session:
        invitation_record = session.scalar(
            select(IdempotencyRecord).where(IdempotencyRecord.request_key == "invite-member10")
        )
        assert invitation_record is not None
        assert "mini_program_code_data_url" not in invitation_record.response_body
    member_headers, _ = _login(client, "dev:member10")
    accepted = client.post(
        "/api/v1/mobile/invitations/accept",
        headers=member_headers,
        json={
            "invitation_token": invitation["invitation_token"],
            "phone_code": "dev:13800000010",
        },
    )

    assert accepted.status_code == 200
    assert accepted.json()["status"] == "bound"
    projects = client.get("/api/v1/mobile/projects", headers=member_headers)
    assert [project["code"] for project in projects.json()] == ["ZPD1322"]
    assert projects.json()[0]["business_date"]
    assert len(projects.json()[0]["milestones"]) == 24
    assert projects.json()[0]["milestones"][0]["assignments"].keys() == {
        "R",
        "A",
        "C",
        "I",
    }
    dashboard = client.get(
        f"/api/v1/mobile/projects/{project_id}/dashboard", headers=member_headers
    )
    assert dashboard.json()["current_version_number"] == 1
    assert dashboard.json()["business_date"]
    assert len(dashboard.json()["milestones"]) == 24

    with client.app.state.session_factory() as session:
        user = session.scalar(select(MobileUser).where(MobileUser.display_name == "dev:member10"))
        binding = session.scalar(
            select(MemberBinding).where(MemberBinding.project_id == uuid.UUID(project_id))
        )
        assert user is not None and binding is not None
        assert user.phone_masked == "138****0010"
        assert user.phone_hash != "13800000010"
        assert user.phone_ciphertext is not None
        assert "13800000010" not in user.phone_ciphertext
        assert user.phone_key_version == 1
        assert binding.provided_phone_masked == "138****0010"
        assert binding.provided_phone_hash != "13800000010"
        assert "13800000010" not in str(user.__dict__)
        assert "13800000010" not in str(binding.__dict__)


def test_invitation_only_binding_requires_the_feature_flag(mobile_workflow: TestClient) -> None:
    client = mobile_workflow
    project_id = _published_project(client)
    invitation = _invite(client, project_id, "成员10", "invite-only-disabled")
    member_headers, _ = _login(client, "dev:invite-only-disabled")

    accepted = client.post(
        "/api/v1/mobile/invitations/accept",
        headers=member_headers,
        json={"invitation_token": invitation["invitation_token"]},
    )

    assert accepted.status_code == 400
    assert accepted.json()["detail"] == "invitation-only binding is disabled"


def test_invitation_only_binding_activates_an_invitation_without_expected_phone(
    mobile_workflow: TestClient,
) -> None:
    client = mobile_workflow
    client.app.state.settings.allow_invitation_only_binding = True
    project_id = _published_project(client)
    invitation = _invite(client, project_id, "成员10", "invite-only-bound")
    member_headers, _ = _login(client, "dev:invite-only-bound")

    accepted = client.post(
        "/api/v1/mobile/invitations/accept",
        headers=member_headers,
        json={"invitation_token": invitation["invitation_token"]},
    )

    assert accepted.status_code == 200
    assert accepted.json()["status"] == "bound"
    assert client.get("/api/v1/mobile/projects", headers=member_headers).json()


def test_invitation_only_binding_with_expected_phone_requires_review(
    mobile_workflow: TestClient,
) -> None:
    client = mobile_workflow
    client.app.state.settings.allow_invitation_only_binding = True
    project_id = _published_project(client)
    invitation = _invite(
        client,
        project_id,
        "成员08",
        "invite-only-review",
        expected_phone="13800000009",
    )
    member_headers, _ = _login(client, "dev:invite-only-review")

    accepted = client.post(
        "/api/v1/mobile/invitations/accept",
        headers=member_headers,
        json={"invitation_token": invitation["invitation_token"]},
    )

    assert accepted.status_code == 200
    assert accepted.json()["status"] == "pending_review"
    assert client.get("/api/v1/mobile/projects", headers=member_headers).json() == []


def test_bound_sheet_project_manager_receives_project_manager_role(
    mobile_workflow: TestClient,
) -> None:
    client = mobile_workflow
    project_id = _published_project(client)
    invitation = _invite(client, project_id, "成员01", "invite-sheet-manager")
    manager_headers, manager = _login(client, "dev:sheet-manager")

    accepted = client.post(
        "/api/v1/mobile/invitations/accept",
        headers=manager_headers,
        json={
            "invitation_token": invitation["invitation_token"],
            "phone": "13800000001",
        },
    )

    assert accepted.status_code == 200
    with client.app.state.session_factory() as session:
        membership = session.scalar(
            select(ProjectMembership).where(
                ProjectMembership.project_id == uuid.UUID(project_id),
                ProjectMembership.actor_id == f"mobile:{manager['user']['id']}",
            )
        )
        assert membership is not None
        assert membership.role == ProjectRole.MANAGER

    responsible = _invite(client, project_id, "成员10", "invite-manager-review-submitter")
    responsible_headers, _ = _login(client, "dev:manager-review-submitter")
    client.post(
        "/api/v1/mobile/invitations/accept",
        headers=responsible_headers,
        json={"invitation_token": responsible["invitation_token"], "phone": "13800000010"},
    )
    proposal = client.post(
        f"/api/v1/mobile/projects/{project_id}/milestones/M23/proposals",
        headers={**responsible_headers, "X-Idempotency-Key": "manager-review-proposal"},
        json={
            "kind": "delay",
            "base_version_number": 1,
            "start_date": "2026-11-16",
            "end_date": "2026-11-20",
            "reason": "请求项目经理审批",
        },
    )
    assert proposal.status_code == 201
    approvable = client.get(
        f"/api/v1/mobile/projects/{project_id}/change-proposals",
        headers=manager_headers,
    )
    assert [item["id"] for item in approvable.json()] == [proposal.json()["id"]]
    projects = client.get("/api/v1/mobile/projects", headers=manager_headers).json()
    assert projects[0]["pending_approval_count"] == 1
    approved = client.post(
        f"/api/v1/mobile/change-proposals/{proposal.json()['id']}/approve",
        headers={**manager_headers, "X-Idempotency-Key": "sheet-manager-approves"},
        json={"expected_project_version": 1},
    )
    assert approved.status_code == 200


def test_reimport_transfers_project_manager_role_without_changing_raci(
    mobile_workflow: TestClient,
) -> None:
    client = mobile_workflow
    project_id = _published_project(client)
    manager_invitation = _invite(client, project_id, "成员01", "invite-old-manager")
    old_headers, old_user = _login(client, "dev:old-manager")
    client.post(
        "/api/v1/mobile/invitations/accept",
        headers=old_headers,
        json={"invitation_token": manager_invitation["invitation_token"], "phone": "13800000001"},
    )
    next_invitation = _invite(client, project_id, "成员02", "invite-new-manager")
    next_headers, next_user = _login(client, "dev:new-manager")
    client.post(
        "/api/v1/mobile/invitations/accept",
        headers=next_headers,
        json={"invitation_token": next_invitation["invitation_token"], "phone": "13800000002"},
    )

    with TemporaryDirectory(dir=ROOT / "tmp") as directory:
        changed = Path(directory) / "new-manager.xlsx"
        shutil.copyfile(WORKBOOK, changed)
        workbook = load_workbook(changed)
        team = workbook["项目团队构成"]
        team["B5"] = "项目统筹"
        team["B6"] = "项目经理"
        workbook.save(changed)
        with changed.open("rb") as source:
            imported = client.post(
                f"/api/v1/projects/{project_id}/imports",
                headers=_admin_headers("import-new-manager"),
                files={
                    "file": (
                        changed.name,
                        source,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
        assert imported.status_code == 201
        published = client.post(
            f"/api/v1/imports/{imported.json()['id']}/publish",
            headers=_admin_headers("publish-new-manager"),
            json={"expected_project_version": 1},
        )
        assert published.status_code == 200

    with client.app.state.session_factory() as session:
        roles = {
            membership.actor_id: membership.role
            for membership in session.scalars(
                select(ProjectMembership).where(
                    ProjectMembership.project_id == uuid.UUID(project_id)
                )
            )
        }
        version = session.scalar(
            select(ProjectVersion).where(
                ProjectVersion.project_id == uuid.UUID(project_id),
                ProjectVersion.version_number == 2,
            )
        )
        assert version is not None
        m01 = next(item for item in version.snapshot["milestones"] if item["code"] == "M01")
        assert m01["assignments"]["R"] == ["成员01"]
        assert m01["assignments"]["A"] == ["成员01"]
        assert roles[f"mobile:{old_user['user']['id']}"] == ProjectRole.ACCOUNTABLE
        assert roles[f"mobile:{next_user['user']['id']}"] == ProjectRole.MANAGER


def test_non_manager_submitter_cannot_resolve_own_ra_proposal(
    mobile_workflow: TestClient,
) -> None:
    client = mobile_workflow
    project_id = _published_project(client)
    invitation = _invite(client, project_id, "成员05", "invite-self-review")
    member_headers, _ = _login(client, "dev:self-review")
    client.post(
        "/api/v1/mobile/invitations/accept",
        headers=member_headers,
        json={"invitation_token": invitation["invitation_token"], "phone": "13800000005"},
    )
    proposal = client.post(
        f"/api/v1/mobile/projects/{project_id}/milestones/M05/proposals",
        headers={**member_headers, "X-Idempotency-Key": "self-review-proposal"},
        json={
            "kind": "completed",
            "base_version_number": 1,
            "actual_completion_date": "2026-08-12",
            "reason": "执行完成，等待独立审批",
        },
    )
    assert proposal.status_code == 201

    visible = client.get(
        f"/api/v1/mobile/projects/{project_id}/change-proposals",
        headers=member_headers,
    )
    assert visible.status_code == 200
    assert visible.json()[0]["id"] == proposal.json()["id"]
    assert visible.json()[0]["is_own_submission"] is True
    assert visible.json()[0]["can_resolve"] is False
    assert visible.json()[0]["created_at"]

    approved = client.post(
        f"/api/v1/mobile/change-proposals/{proposal.json()['id']}/approve",
        headers={**member_headers, "X-Idempotency-Key": "self-review-approve"},
        json={"expected_project_version": 1},
    )
    rejected = client.post(
        f"/api/v1/mobile/change-proposals/{proposal.json()['id']}/reject",
        headers={**member_headers, "X-Idempotency-Key": "self-review-reject"},
        json={"reason": "不能自审"},
    )

    assert approved.status_code == 403
    assert rejected.status_code == 403
    manager_approved = client.post(
        f"/api/v1/change-proposals/{proposal.json()['id']}/approve",
        headers=_admin_headers("manager-approves-ra-proposal"),
        json={"expected_project_version": 1},
    )
    assert manager_approved.status_code == 200


def test_sheet_project_manager_can_approve_own_proposal(
    mobile_workflow: TestClient,
) -> None:
    client = mobile_workflow
    project_id = _published_project(client)
    invitation = _invite(client, project_id, "成员01", "invite-manager-self-review")
    manager_headers, _ = _login(client, "dev:manager-self-review")
    client.post(
        "/api/v1/mobile/invitations/accept",
        headers=manager_headers,
        json={"invitation_token": invitation["invitation_token"], "phone": "13800000001"},
    )
    proposal = client.post(
        f"/api/v1/mobile/projects/{project_id}/milestones/M01/proposals",
        headers={**manager_headers, "X-Idempotency-Key": "manager-self-proposal"},
        json={
            "kind": "completed",
            "base_version_number": 1,
            "actual_completion_date": "2026-08-12",
            "reason": "项目经理本人提交并审批",
        },
    )
    assert proposal.status_code == 201
    visible = client.get(
        f"/api/v1/mobile/projects/{project_id}/change-proposals",
        headers=manager_headers,
    )
    assert visible.status_code == 200
    assert visible.json()[0]["is_own_submission"] is True
    assert visible.json()[0]["can_resolve"] is True

    approved = client.post(
        f"/api/v1/mobile/change-proposals/{proposal.json()['id']}/approve",
        headers={**manager_headers, "X-Idempotency-Key": "manager-self-approve"},
        json={"expected_project_version": 1},
    )
    assert approved.status_code == 200


def test_my_tasks_only_returns_bound_members_ra_milestones_without_duplicates(
    mobile_workflow: TestClient,
) -> None:
    client = mobile_workflow
    project_id = _published_project(client)
    outsider_headers, _ = _login(client, "dev:my-tasks-outsider")

    assert client.get("/api/v1/mobile/my-tasks", headers=outsider_headers).json() == []

    invitation = _invite(client, project_id, "成员03", "invite-my-tasks-member03")
    member_headers, _ = _login(client, "dev:my-tasks-member03")
    accepted = client.post(
        "/api/v1/mobile/invitations/accept",
        headers=member_headers,
        json={
            "invitation_token": invitation["invitation_token"],
            "phone_code": "dev:13800000003",
        },
    )
    assert accepted.status_code == 200

    with client.app.state.session_factory() as session:
        version = session.scalar(
            select(ProjectVersion).where(ProjectVersion.project_id == uuid.UUID(project_id))
        )
        assert version is not None
        snapshot = dict(version.snapshot)
        milestones = [dict(item) for item in snapshot["milestones"]]
        target = milestones[0]
        assignments = {key: list(value) for key, value in target["assignments"].items()}
        assignments["R"] = ["成员03"]
        assignments["A"] = ["成员03"]
        target["assignments"] = assignments
        snapshot["milestones"] = milestones
        version.snapshot = snapshot
        session.commit()

    response = client.get("/api/v1/mobile/my-tasks", headers=member_headers)
    assert response.status_code == 200
    projects = response.json()
    assert len(projects) == 1
    assert projects[0]["project"] == {
        "id": project_id,
        "code": "ZPD1322",
        "name": "Lyra Pro",
    }
    assert projects[0]["business_date"]
    assert projects[0]["member_name"] == "成员03"
    matching = [task for task in projects[0]["tasks"] if task["code"] == target["code"]]
    assert len(matching) == 1
    assert matching[0]["roles"] == ["R", "A"]
    assert matching[0]["risk"] == "overdue"
    assert all(set(task["roles"]) <= {"R", "A"} for task in projects[0]["tasks"])


def test_bound_user_can_view_read_only_project_review_without_contacts(
    mobile_workflow: TestClient,
) -> None:
    client = mobile_workflow
    project_id = _published_project(client)
    outsider_headers, _ = _login(client, "dev:review-outsider")
    forbidden = client.get(
        f"/api/v1/mobile/projects/{project_id}/review",
        headers=outsider_headers,
    )
    assert forbidden.status_code == 403

    invitation = _invite(client, project_id, "成员10", "invite-review-member10")
    member_headers, _ = _login(client, "dev:review-member10")
    accepted = client.post(
        "/api/v1/mobile/invitations/accept",
        headers=member_headers,
        json={
            "invitation_token": invitation["invitation_token"],
            "phone_code": "dev:13800000010",
        },
    )
    assert accepted.status_code == 200

    review = client.get(
        f"/api/v1/mobile/projects/{project_id}/review",
        headers=member_headers,
    )

    assert review.status_code == 200
    payload = review.json()
    assert len(payload["product_specs"]) == 70
    assert "check_confirmation" not in payload["product_specs"][0]
    assert "check_content" not in payload["product_specs"][0]
    assert len(payload["members"]) == 22
    assert len(payload["milestones"]) == 24
    assert payload["milestones"][0]["assignments"].keys() == {"R", "A", "C", "I"}
    assert "phone" not in payload["members"][0]
    assert "email" not in payload["members"][0]


def test_mobile_token_cannot_use_admin_project_data_change_endpoints(
    mobile_workflow: TestClient,
) -> None:
    client = mobile_workflow
    project_id = _published_project(client)
    member_headers, _ = _login(client, "dev:project-data-boundary")

    editable = client.get(
        f"/api/v1/projects/{project_id}/editable-data",
        headers=member_headers,
    )
    assert editable.status_code == 403

    create_change_set = client.post(
        f"/api/v1/projects/{project_id}/change-sets",
        headers={**member_headers, "X-Idempotency-Key": "mobile-admin-boundary"},
        json={
            "base_version_number": 1,
            "source": "mini_program",
            "reason": "尝试越权修改项目基础数据",
            "operations": [
                {
                    "op": "replace",
                    "resource": "product_spec",
                    "key": "OS版本",
                    "value": {"item": "OS版本", "value": "Android 17"},
                }
            ],
        },
    )
    assert create_change_set.status_code == 403


def test_user_can_register_an_explicit_wechat_subscription_grant(
    mobile_workflow: TestClient,
) -> None:
    headers, _ = _login(mobile_workflow, "dev:subscriber")

    first = mobile_workflow.post(
        "/api/v1/mobile/subscription-grants",
        headers=headers,
        json={"template_id": "test-template"},
    )
    second = mobile_workflow.post(
        "/api/v1/mobile/subscription-grants",
        headers=headers,
        json={"template_id": "test-template"},
    )

    assert first.status_code == 200
    assert second.json() == {"template_id": "test-template", "remaining_uses": 2}


def test_phone_mismatch_requires_manager_review(mobile_workflow: TestClient) -> None:
    client = mobile_workflow
    project_id = _published_project(client)
    invitation = _invite(
        client,
        project_id,
        "成员08",
        "invite-member08",
        expected_phone="13800000008",
    )
    member_headers, _ = _login(client, "dev:member08")

    accepted = client.post(
        "/api/v1/mobile/invitations/accept",
        headers=member_headers,
        json={"invitation_token": invitation["invitation_token"], "phone": "13900000008"},
    )
    assert accepted.status_code == 200
    assert accepted.json()["status"] == "pending_review"
    assert client.get("/api/v1/mobile/projects", headers=member_headers).json() == []
    pending = client.get(f"/api/v1/projects/{project_id}/member-bindings", headers=PM)
    assert pending.status_code == 200
    assert pending.json()[0]["member_name"] == "成员08"
    assert pending.json()[0]["status"] == "pending_review"
    assert pending.json()[0]["expected_phone"] == "138****0008"
    assert pending.json()[0]["provided_phone"] == "139****0008"

    approved = client.post(
        f"/api/v1/member-bindings/{accepted.json()['id']}/approve",
        headers=_admin_headers("approve-member08"),
    )
    assert approved.status_code == 200
    assert approved.json()["status"] == "bound"
    assert len(client.get("/api/v1/mobile/projects", headers=member_headers).json()) == 1


def test_production_binding_rejects_a_phone_number_supplied_by_the_client(
    mobile_workflow: TestClient,
) -> None:
    client = mobile_workflow
    project_id = _published_project(client)
    invitation = _invite(client, project_id, "成员10", "invite-production-phone")
    member_headers, _ = _login(client, "dev:production-phone")
    client.app.state.settings.allow_development_wechat_login = False

    accepted = client.post(
        "/api/v1/mobile/invitations/accept",
        headers=member_headers,
        json={"invitation_token": invitation["invitation_token"], "phone": "13800000010"},
    )

    assert accepted.status_code == 400
    assert accepted.json()["detail"] == "direct phone input is allowed only in development"


def test_same_mobile_user_cannot_bind_two_member_identities_in_one_project(
    mobile_workflow: TestClient,
) -> None:
    client = mobile_workflow
    project_id = _published_project(client)
    first = _invite(client, project_id, "成员10", "invite-first-identity")
    second = _invite(client, project_id, "成员09", "invite-second-identity")
    headers, _ = _login(client, "dev:duplicate-identity")
    accepted = client.post(
        "/api/v1/mobile/invitations/accept",
        headers=headers,
        json={"invitation_token": first["invitation_token"], "phone": "13800000010"},
    )
    assert accepted.status_code == 200

    duplicate = client.post(
        "/api/v1/mobile/invitations/accept",
        headers=headers,
        json={"invitation_token": second["invitation_token"], "phone": "13800000009"},
    )

    assert duplicate.status_code == 409


def test_responsible_member_can_submit_own_milestone_but_not_another(
    mobile_workflow: TestClient,
) -> None:
    client = mobile_workflow
    project_id = _published_project(client)
    invitation = _invite(client, project_id, "成员10", "invite-progress-member10")
    member_headers, _ = _login(client, "dev:progress-member10")
    client.post(
        "/api/v1/mobile/invitations/accept",
        headers=member_headers,
        json={"invitation_token": invitation["invitation_token"], "phone": "13800000010"},
    )

    own = client.post(
        f"/api/v1/mobile/projects/{project_id}/milestones/M23/proposals",
        headers={**member_headers, "X-Idempotency-Key": "mobile-complete-m23"},
        json={
            "kind": "completed",
            "base_version_number": 1,
            "actual_completion_date": "2026-08-06",
            "reason": "软件封板已完成",
        },
    )
    other = client.post(
        f"/api/v1/mobile/projects/{project_id}/milestones/M01/proposals",
        headers={**member_headers, "X-Idempotency-Key": "mobile-complete-m01"},
        json={
            "kind": "completed",
            "base_version_number": 1,
            "actual_completion_date": "2026-08-06",
            "reason": "无权更新的节点",
        },
    )

    assert own.status_code == 201
    assert other.status_code == 403
    accountable_invitation = _invite(client, project_id, "成员09", "invite-accountable-member09")
    accountable_headers, _ = _login(client, "dev:accountable-member09")
    client.post(
        "/api/v1/mobile/invitations/accept",
        headers=accountable_headers,
        json={
            "invitation_token": accountable_invitation["invitation_token"],
            "phone": "13800000009",
        },
    )
    approved = client.post(
        f"/api/v1/mobile/change-proposals/{own.json()['id']}/approve",
        headers={**accountable_headers, "X-Idempotency-Key": "approve-mobile-m23"},
        json={"expected_project_version": 1},
    )
    assert approved.status_code == 200
    assert approved.json()["version_number"] == 2
    dashboard = client.get(
        f"/api/v1/mobile/projects/{project_id}/dashboard", headers=member_headers
    ).json()
    milestone = next(item for item in dashboard["milestones"] if item["code"] == "M23")
    assert milestone["actual_completion"]["end_date"] == "2026-08-06"


def test_accountable_member_can_list_only_approvable_pending_proposals(
    mobile_workflow: TestClient,
) -> None:
    client = mobile_workflow
    project_id = _published_project(client)
    responsible = _invite(client, project_id, "成员10", "invite-r-list")
    responsible_headers, _ = _login(client, "dev:r-list")
    client.post(
        "/api/v1/mobile/invitations/accept",
        headers=responsible_headers,
        json={"invitation_token": responsible["invitation_token"], "phone": "13800000010"},
    )
    proposal = client.post(
        f"/api/v1/mobile/projects/{project_id}/milestones/M23/proposals",
        headers={**responsible_headers, "X-Idempotency-Key": "proposal-for-a-list"},
        json={
            "kind": "delay",
            "base_version_number": 1,
            "start_date": "2026-11-16",
            "end_date": "2026-11-20",
            "reason": "联调延期",
        },
    ).json()
    accountable = _invite(client, project_id, "成员09", "invite-a-list")
    accountable_headers, _ = _login(client, "dev:a-list")
    client.post(
        "/api/v1/mobile/invitations/accept",
        headers=accountable_headers,
        json={"invitation_token": accountable["invitation_token"], "phone": "13800000009"},
    )

    response = client.get(
        f"/api/v1/mobile/projects/{project_id}/change-proposals",
        headers=accountable_headers,
    )

    assert response.status_code == 200
    assert [item["id"] for item in response.json()] == [proposal["id"]]


def test_only_accountable_member_can_reject_mobile_proposal(
    mobile_workflow: TestClient,
) -> None:
    client = mobile_workflow
    project_id = _published_project(client)
    responsible = _invite(client, project_id, "成员10", "invite-r-reject")
    responsible_headers, _ = _login(client, "dev:r-reject")
    client.post(
        "/api/v1/mobile/invitations/accept",
        headers=responsible_headers,
        json={"invitation_token": responsible["invitation_token"], "phone": "13800000010"},
    )
    proposal = client.post(
        f"/api/v1/mobile/projects/{project_id}/milestones/M23/proposals",
        headers={**responsible_headers, "X-Idempotency-Key": "proposal-for-a-reject"},
        json={
            "kind": "delay",
            "base_version_number": 1,
            "start_date": "2026-11-16",
            "end_date": "2026-11-20",
            "reason": "联调延期",
        },
    ).json()
    forbidden = client.post(
        f"/api/v1/mobile/change-proposals/{proposal['id']}/reject",
        headers={**responsible_headers, "X-Idempotency-Key": "r-cannot-reject"},
        json={"reason": "执行者不能驳回"},
    )
    assert forbidden.status_code == 403

    accountable = _invite(client, project_id, "成员09", "invite-a-reject")
    accountable_headers, _ = _login(client, "dev:a-reject")
    client.post(
        "/api/v1/mobile/invitations/accept",
        headers=accountable_headers,
        json={"invitation_token": accountable["invitation_token"], "phone": "13800000009"},
    )
    rejected = client.post(
        f"/api/v1/mobile/change-proposals/{proposal['id']}/reject",
        headers={**accountable_headers, "X-Idempotency-Key": "a-rejects-proposal"},
        json={"reason": "资源未落实，请重新计划"},
    )

    assert rejected.status_code == 200
    assert rejected.json()["status"] == "rejected"
    dashboard = client.get(f"/api/v1/projects/{project_id}/dashboard", headers=PM)
    assert dashboard.json()["current_version_number"] == 1


def test_mobile_issue_message_center_and_natural_language_prefill(
    mobile_workflow: TestClient,
) -> None:
    client = mobile_workflow
    project_id = _published_project(client)
    invitation = _invite(client, project_id, "成员10", "invite-mobile-features")
    member_headers, _ = _login(client, "dev:mobile-features")
    client.post(
        "/api/v1/mobile/invitations/accept",
        headers=member_headers,
        json={"invitation_token": invitation["invitation_token"], "phone": "13800000010"},
    )

    proposal = client.post(
        f"/api/v1/mobile/projects/{project_id}/issue-create-proposals",
        headers={**member_headers, "X-Idempotency-Key": "mobile-issue"},
        json={
            "description": "软件封板存在阻塞",
            "impact": "影响PVT",
            "owner_name": "成员10",
            "accountable_names": ["成员02"],
            "severity": "high",
            "due_date": "2026-08-20",
        },
    )
    assert proposal.status_code == 201
    approved = client.post(
        f"/api/v1/issue-create-proposals/{proposal.json()['id']}/approve",
        headers=_admin_headers("mobile-issue-approve"),
    )
    assert approved.status_code == 200
    issue_id = approved.json()["issue_id"]
    client.get(
        f"/api/v1/mobile/projects/{project_id}/issues", headers=member_headers
    ).json()[0]
    assigned_owner = client.post(
        f"/api/v1/mobile/projects/{project_id}/issue-create-proposals",
        headers={**member_headers, "X-Idempotency-Key": "mobile-forged-owner"},
        json={
            "description": "伪造责任人",
            "impact": "错误通知",
            "owner_name": "成员11",
            "accountable_names": ["成员02"],
            "severity": "high",
            "due_date": "2026-08-21",
        },
    )
    assert assigned_owner.status_code == 201
    assert assigned_owner.json()["payload"]["owner_name"] == "成员11"
    updated = client.patch(
        f"/api/v1/mobile/issues/{issue_id}",
        headers={**member_headers, "X-Idempotency-Key": "mobile-issue-update"},
        json={"expected_revision": 1, "status": "处理中"},
    )
    assert updated.status_code == 200
    assert updated.json()["revision"] == 2
    raci_update = client.patch(
        f"/api/v1/mobile/issues/{issue_id}",
        headers={**member_headers, "X-Idempotency-Key": "mobile-issue-raci-update"},
        json={"expected_revision": 2, "accountable_names": ["成员03"]},
    )
    assert raci_update.status_code == 200
    assert raci_update.json()["accountable_names"] == ["成员03"]
    other_invitation = _invite(client, project_id, "成员11", "invite-other-issue-owner")
    other_headers, _ = _login(client, "dev:other-issue-owner")
    client.post(
        "/api/v1/mobile/invitations/accept",
        headers=other_headers,
        json={"invitation_token": other_invitation["invitation_token"], "phone": "13800000011"},
    )
    forbidden_delete = client.request(
        "DELETE",
        f"/api/v1/mobile/issues/{issue_id}",
        headers={**other_headers, "X-Idempotency-Key": "mobile-issue-delete-forbidden"},
        json={"expected_revision": 3, "reason": "越权作废"},
    )
    assert forbidden_delete.status_code == 403
    requested_delete = client.request(
        "DELETE",
        f"/api/v1/mobile/issues/{issue_id}",
        headers={**member_headers, "X-Idempotency-Key": "mobile-issue-delete"},
        json={"expected_revision": 3, "reason": "问题已作废"},
    )
    assert requested_delete.status_code == 201
    assert requested_delete.json()["status"] == "pending"
    active_issue = client.get(
        f"/api/v1/mobile/projects/{project_id}/issues", headers=member_headers
    ).json()[0]
    assert active_issue["revision"] == 3
    manager_invitation = _invite(client, project_id, "成员01", "invite-delete-manager")
    manager_headers, _ = _login(client, "dev:delete-manager")
    client.post(
        "/api/v1/mobile/invitations/accept",
        headers=manager_headers,
        json={"invitation_token": manager_invitation["invitation_token"], "phone": "13800000001"},
    )
    manager_queue = client.get(
        f"/api/v1/mobile/projects/{project_id}/issue-delete-proposals",
        headers=manager_headers,
    )
    assert [item["id"] for item in manager_queue.json()] == [requested_delete.json()["id"]]
    approved_delete = client.post(
        f"/api/v1/mobile/issue-delete-proposals/{requested_delete.json()['id']}/approve",
        headers={**manager_headers, "X-Idempotency-Key": "mobile-issue-delete-approve"},
    )
    assert approved_delete.status_code == 200
    assert approved_delete.json()["status"] == "approved"
    approval_messages = client.get("/api/v1/mobile/messages", headers=member_headers).json()
    assert any(item["type"] == "issue_delete_approved" for item in approval_messages)

    prefill = client.post(
        "/api/v1/mobile/natural-language/prefill",
        headers=member_headers,
        json={"text": "M23延期到2026-08-30，原因是驱动联调"},
    )
    assert prefill.status_code == 200
    assert prefill.json()["milestone_code"] == "M23"
    assert prefill.json()["end_date"] == "2026-08-30"
    assert prefill.json()["requires_confirmation"] is True

    messages = client.get("/api/v1/mobile/messages", headers=member_headers)
    assert messages.status_code == 200
    assert any(item["type"] == "binding_approved" for item in messages.json())
    assert all(item["created_at"].endswith("+00:00") for item in messages.json())
    message = messages.json()[0]
    marked = client.patch(
        f"/api/v1/mobile/messages/{message['id']}/read",
        headers={**member_headers, "X-Idempotency-Key": "read-binding-message"},
    )
    assert marked.status_code == 200
    assert marked.json()["is_read"] is True


def test_mobile_issue_create_requires_manager_approval_and_allows_any_member_r(
    mobile_workflow: TestClient,
) -> None:
    client = mobile_workflow
    project_id = _published_project(client)
    submitter_invitation = _invite(client, project_id, "成员10", "invite-issue-submitter")
    submitter_headers, _ = _login(client, "dev:issue-submitter")
    client.post(
        "/api/v1/mobile/invitations/accept",
        headers=submitter_headers,
        json={"invitation_token": submitter_invitation["invitation_token"], "phone": "13800000010"},
    )
    manager_invitation = _invite(client, project_id, "成员01", "invite-issue-manager")
    manager_headers, _ = _login(client, "dev:issue-manager")
    client.post(
        "/api/v1/mobile/invitations/accept",
        headers=manager_headers,
        json={"invitation_token": manager_invitation["invitation_token"], "phone": "13800000001"},
    )

    invalid_owner = client.post(
        f"/api/v1/mobile/projects/{project_id}/issue-create-proposals",
        headers={**submitter_headers, "X-Idempotency-Key": "invalid-issue-owner"},
        json={
            "description": "责任人不在团队",
            "impact": "无法落实责任",
            "owner_name": "外部人员",
            "accountable_names": ["成员02"],
            "severity": "high",
            "due_date": "2026-08-20",
        },
    )
    assert invalid_owner.status_code == 409

    proposed = client.post(
        f"/api/v1/mobile/projects/{project_id}/issue-create-proposals",
        headers={**submitter_headers, "X-Idempotency-Key": "issue-create-proposal"},
        json={
            "description": "张三反馈、李四负责的问题",
            "impact": "影响试产",
            "owner_name": "成员11",
            "accountable_names": ["成员02"],
            "consulted_names": ["成员03"],
            "informed_names": ["成员04"],
            "severity": "high",
            "due_date": "2026-08-20",
        },
    )
    assert proposed.status_code == 201, proposed.json()
    assert proposed.json()["status"] == "pending"
    assert proposed.json()["payload"]["owner_name"] == "成员11"
    assert client.get(
        f"/api/v1/mobile/projects/{project_id}/issues", headers=submitter_headers
    ).json() == []
    assert client.get(
        f"/api/v1/mobile/projects/{project_id}/issue-create-proposals",
        headers=submitter_headers,
    ).status_code == 403
    assert client.post(
        f"/api/v1/mobile/issue-create-proposals/{proposed.json()['id']}/approve",
        headers={**submitter_headers, "X-Idempotency-Key": "forbidden-approve-issue-create"},
    ).status_code == 403
    dashboard = client.get(f"/api/v1/projects/{project_id}/dashboard", headers=PM)
    assert dashboard.json()["counts"]["issues_open"] == 0

    manager_queue = client.get(
        f"/api/v1/mobile/projects/{project_id}/issue-create-proposals",
        headers=manager_headers,
    )
    assert [item["id"] for item in manager_queue.json()] == [proposed.json()["id"]]
    approved = client.post(
        f"/api/v1/mobile/issue-create-proposals/{proposed.json()['id']}/approve",
        headers={**manager_headers, "X-Idempotency-Key": "approve-issue-create"},
    )
    assert approved.status_code == 200, approved.json()
    assert approved.json()["status"] == "approved"
    assert approved.json()["issue_id"]
    assert client.get(
        f"/api/v1/mobile/projects/{project_id}/issue-create-proposals",
        headers=manager_headers,
    ).json() == []
    issues = client.get(
        f"/api/v1/mobile/projects/{project_id}/issues", headers=submitter_headers
    ).json()
    assert len(issues) == 1
    assert issues[0]["owner_name"] == "成员11"
    messages = client.get("/api/v1/mobile/messages", headers=submitter_headers).json()
    assert any(item["type"] == "issue_create_approved" for item in messages)
    audit = client.get(f"/api/v1/projects/{project_id}/audit-logs", headers=PM).json()
    assert "issue_create_proposal.created" in [item["action"] for item in audit]
    assert "issue_create_proposal.approved" in [item["action"] for item in audit]


def test_project_manager_can_reject_issue_create_proposal(
    mobile_workflow: TestClient,
) -> None:
    client = mobile_workflow
    project_id = _published_project(client)
    submitter_invitation = _invite(client, project_id, "成员10", "invite-issue-reject-submitter")
    submitter_headers, _ = _login(client, "dev:issue-reject-submitter")
    client.post(
        "/api/v1/mobile/invitations/accept",
        headers=submitter_headers,
        json={"invitation_token": submitter_invitation["invitation_token"], "phone": "13800000010"},
    )
    proposed = client.post(
        f"/api/v1/mobile/projects/{project_id}/issue-create-proposals",
        headers={**submitter_headers, "X-Idempotency-Key": "issue-create-reject-proposal"},
        json={
            "description": "不满足重点问题标准",
            "impact": "影响较小",
            "owner_name": "成员11",
            "accountable_names": ["成员02"],
            "severity": "low",
            "due_date": "2026-08-20",
        },
    )
    rejected = client.post(
        f"/api/v1/issue-create-proposals/{proposed.json()['id']}/reject",
        headers=_admin_headers("reject-issue-create"),
        json={"reason": "不满足重点问题标准"},
    )
    assert rejected.status_code == 200
    assert rejected.json()["status"] == "rejected"
    assert rejected.json()["issue_id"] is None
    messages = client.get("/api/v1/mobile/messages", headers=submitter_headers).json()
    assert any(item["type"] == "issue_create_rejected" for item in messages)
    assert client.get(f"/api/v1/projects/{project_id}/issues", headers=PM).json() == []


def test_project_manager_rejects_issue_delete_and_notifies_submitter(
    mobile_workflow: TestClient,
) -> None:
    client = mobile_workflow
    project_id = _published_project(client)
    owner_invitation = _invite(client, project_id, "成员10", "invite-delete-reject-owner")
    owner_headers, _ = _login(client, "dev:delete-reject-owner")
    client.post(
        "/api/v1/mobile/invitations/accept",
        headers=owner_headers,
        json={"invitation_token": owner_invitation["invitation_token"], "phone": "13800000010"},
    )
    manager_invitation = _invite(client, project_id, "成员01", "invite-delete-reject-manager")
    manager_headers, _ = _login(client, "dev:delete-reject-manager")
    client.post(
        "/api/v1/mobile/invitations/accept",
        headers=manager_headers,
        json={"invitation_token": manager_invitation["invitation_token"], "phone": "13800000001"},
    )
    created = client.post(
        f"/api/v1/mobile/projects/{project_id}/issue-create-proposals",
        headers={**owner_headers, "X-Idempotency-Key": "delete-reject-create"},
        json={
            "description": "需要继续跟踪的问题",
            "impact": "影响试产",
            "owner_name": "成员10",
            "accountable_names": ["成员02"],
            "severity": "high",
            "due_date": "2026-08-20",
        },
    )
    approved = client.post(
        f"/api/v1/mobile/issue-create-proposals/{created.json()['id']}/approve",
        headers={**manager_headers, "X-Idempotency-Key": "delete-reject-create-approve"},
    )
    updated = client.patch(
        f"/api/v1/mobile/issues/{approved.json()['issue_id']}",
        headers={**owner_headers, "X-Idempotency-Key": "update-issue-raci"},
        json={
            "expected_revision": 1,
            "accountable_names": ["成员03"],
            "consulted_names": ["成员04"],
            "informed_names": ["成员05"],
        },
    )
    assert updated.status_code == 200, updated.json()
    assert updated.json()["accountable_names"] == ["成员03"]
    my_tasks = client.get("/api/v1/mobile/my-tasks", headers=owner_headers).json()
    issue_tasks = [
        task
        for project in my_tasks
        for task in project["tasks"]
        if task["kind"] == "issue"
    ]
    assert issue_tasks[0]["name"] == "需要继续跟踪的问题"
    assert issue_tasks[0]["roles"] == ["R"]
    requested = client.request(
        "DELETE",
        f"/api/v1/mobile/issues/{approved.json()['issue_id']}",
        headers={**owner_headers, "X-Idempotency-Key": "delete-reject-request"},
        json={"expected_revision": 2, "reason": "误判为重复"},
    )
    rejected = client.post(
        f"/api/v1/mobile/issue-delete-proposals/{requested.json()['id']}/reject",
        headers={**manager_headers, "X-Idempotency-Key": "delete-reject-resolve"},
        json={"reason": "仍需继续跟踪"},
    )

    assert rejected.status_code == 200
    assert rejected.json()["status"] == "rejected"
    messages = client.get("/api/v1/mobile/messages", headers=owner_headers).json()
    assert any(item["type"] == "issue_delete_rejected" for item in messages)
    issue = client.get(
        f"/api/v1/mobile/projects/{project_id}/issues", headers=owner_headers
    ).json()[0]
    assert issue["status"] != "已关闭"
    assert issue["revision"] == 2
    reassigned = client.patch(
        f"/api/v1/mobile/issues/{approved.json()['issue_id']}",
        headers={**owner_headers, "X-Idempotency-Key": "reassign-issue-owner"},
        json={"expected_revision": 2, "owner_name": "成员11"},
    )
    assert reassigned.status_code == 200, reassigned.json()
    assert reassigned.json()["owner_name"] == "成员11"


def test_publishing_team_change_revokes_removed_member_access(
    mobile_workflow: TestClient,
) -> None:
    client = mobile_workflow
    project_id = _published_project(client)
    invitation = _invite(client, project_id, "成员10", "invite-revoked-member")
    unused_invitation = _invite(client, project_id, "成员11", "invite-revoked-unused-member")
    member_headers, _ = _login(client, "dev:revoked-member")
    client.post(
        "/api/v1/mobile/invitations/accept",
        headers=member_headers,
        json={"invitation_token": invitation["invitation_token"], "phone": "13800000010"},
    )
    with TemporaryDirectory(dir=ROOT / "tmp") as directory:
        changed = Path(directory) / "team-changed.xlsx"
        shutil.copyfile(WORKBOOK, changed)
        workbook = load_workbook(changed)
        team = workbook["项目团队构成"]
        for row in range(5, 27):
            if team[f"C{row}"].value == "成员10":
                team[f"C{row}"] = "成员23"
            if team[f"C{row}"].value == "成员11":
                team[f"C{row}"] = "成员24"
        workbook.save(changed)
        workbook.close()
        with changed.open("rb") as source:
            imported = client.post(
                f"/api/v1/projects/{project_id}/imports",
                headers=_admin_headers("import-team-change"),
                files={
                    "file": (
                        changed.name,
                        source,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            ).json()
    published = client.post(
        f"/api/v1/imports/{imported['id']}/publish",
        headers=_admin_headers("publish-team-change"),
        json={"expected_project_version": 1},
    )
    assert published.status_code == 200

    assert client.get("/api/v1/mobile/projects", headers=member_headers).json() == []
    dashboard = client.get(
        f"/api/v1/mobile/projects/{project_id}/dashboard", headers=member_headers
    )
    assert dashboard.status_code == 403
    unused_headers, _ = _login(client, "dev:revoked-unused-member")
    revoked_accept = client.post(
        "/api/v1/mobile/invitations/accept",
        headers=unused_headers,
        json={
            "invitation_token": unused_invitation["invitation_token"],
            "phone": "13800000011",
        },
    )
    assert revoked_accept.status_code == 404


def test_manager_can_reject_mobile_proposal_without_publishing(
    mobile_workflow: TestClient,
) -> None:
    client = mobile_workflow
    project_id = _published_project(client)
    invitation = _invite(client, project_id, "成员10", "invite-rejected-proposal")
    member_headers, _ = _login(client, "dev:rejected-proposal")
    client.post(
        "/api/v1/mobile/invitations/accept",
        headers=member_headers,
        json={"invitation_token": invitation["invitation_token"], "phone": "13800000010"},
    )
    proposal = client.post(
        f"/api/v1/mobile/projects/{project_id}/milestones/M23/proposals",
        headers={**member_headers, "X-Idempotency-Key": "rejected-m23"},
        json={
            "kind": "delay",
            "base_version_number": 1,
            "start_date": "2026-08-30",
            "end_date": "2026-09-01",
            "reason": "驱动联调延期",
        },
    ).json()

    pending = client.get(f"/api/v1/projects/{project_id}/change-proposals", headers=PM)
    assert pending.status_code == 200
    assert pending.json()[0]["status"] == "pending"
    rejected = client.post(
        f"/api/v1/change-proposals/{proposal['id']}/reject",
        headers=_admin_headers("reject-m23"),
        json={"reason": "不接受延期，请调整资源"},
    )

    assert rejected.status_code == 200
    assert rejected.json()["status"] == "rejected"
    dashboard = client.get(f"/api/v1/projects/{project_id}/dashboard", headers=PM)
    assert dashboard.json()["current_version_number"] == 1
