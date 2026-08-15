import json
import shutil
import zipfile
from collections.abc import Iterator
from datetime import date
from pathlib import Path
from tempfile import TemporaryDirectory

import pytest
from openpyxl import load_workbook

from project_manager_api.domain.models import PlanDateState, PlanWindow, RaciRole
from project_manager_api.imports.errors import (
    InvalidWorkbookError,
    TemplateNotSupportedError,
    WorkbookValidationError,
)
from project_manager_api.imports.registry import ParserRegistry

ROOT = Path(__file__).resolve().parents[3]
FIXTURE_DIR = ROOT / "tests" / "fixtures" / "lyra-template-v1"
WORKBOOK = FIXTURE_DIR / "lyra_v1_sanitized.xlsx"
EXPECTED = json.loads((FIXTURE_DIR / "expected.json").read_text(encoding="utf-8"))
MANIFEST = ROOT / "config" / "templates" / "lyra_project_spec-v1.0.yaml"


@pytest.fixture
def registry() -> ParserRegistry:
    return ParserRegistry.from_manifest_paths([MANIFEST])


@pytest.fixture
def phase1_tmp_path() -> Iterator[Path]:
    with TemporaryDirectory(dir=ROOT / "tmp") as directory:
        yield Path(directory)


def test_xlsx_01_parses_lyra_baseline_to_canonical_draft(registry: ParserRegistry) -> None:
    result = registry.parse(WORKBOOK)

    assert result.draft.template_id == EXPECTED["template"]["id"]
    assert result.draft.template_version == EXPECTED["template"]["version"]
    assert result.draft.document_version == "V1.3"
    assert result.draft.project.code == EXPECTED["project"]["code"]
    assert result.draft.project.name == EXPECTED["project"]["name"]
    assert len(result.draft.members) == 22
    assert len(result.draft.milestones) == 24
    assert len(result.draft.product_specs) == 70
    assert [member.role for member in result.draft.members] == EXPECTED["team_roles"]
    assert [milestone.name for milestone in result.draft.milestones] == EXPECTED["milestones"]
    specs_by_item: dict[str, str | None] = {}
    for item in result.draft.product_specs:
        specs_by_item.setdefault(item.item, item.configuration)
    assert {name: specs_by_item[name] for name in EXPECTED["product_spec_spot_checks"]} == EXPECTED[
        "product_spec_spot_checks"
    ]
    assert result.report.errors == []


def test_xlsx_02_uses_latest_non_empty_plan_snapshot_as_candidate_current_plan(
    registry: ParserRegistry,
) -> None:
    result = registry.parse(WORKBOOK)

    assert [plan.name for plan in result.draft.plan_versions] == [
        "原计划(不改板-库存500个PCB)",
        "变更计划1（改板）",
        "变更计划2（两次试产）",
    ]
    assert result.draft.active_plan_name == EXPECTED["active_plan_name"]
    assert len(result.draft.active_plan.milestones) == 24
    actual_plan = {
        name: _serialize_plan_window(window)
        for name, window in result.draft.active_plan.milestones.items()
    }
    assert actual_plan == EXPECTED["active_plan"]


@pytest.mark.parametrize(
    ("blank_rows", "expected_active_plan"),
    [
        ([5], "变更计划1（改板）"),
        ([4, 5], "原计划(不改板-库存500个PCB)"),
    ],
)
def test_xlsx_plan_selection_falls_back_when_later_change_plan_is_empty(
    registry: ParserRegistry,
    phase1_tmp_path: Path,
    blank_rows: list[int],
    expected_active_plan: str,
) -> None:
    changed = phase1_tmp_path / "empty-later-plan.xlsx"
    shutil.copyfile(WORKBOOK, changed)
    workbook = load_workbook(changed)
    sheet = workbook["项目进度表"]
    for row_number in blank_rows:
        for column in range(2, 26):
            sheet.cell(row_number, column).value = None
    workbook.save(changed)
    workbook.close()

    result = registry.parse(changed)

    assert result.draft.active_plan_name == expected_active_plan
    assert any(
        window.state is not PlanDateState.TBD
        for window in result.draft.active_plan.milestones.values()
    )


def test_xlsx_03_distinguishes_blank_from_not_applicable(registry: ParserRegistry) -> None:
    active_plan = registry.parse(WORKBOOK).draft.active_plan

    assert active_plan.milestones["DVT发板"].state is PlanDateState.NOT_APPLICABLE
    assert active_plan.milestones["DVT-SMT贴片"].state is PlanDateState.TBD


def test_xlsx_04_parses_single_dates_and_date_ranges(registry: ParserRegistry) -> None:
    active_plan = registry.parse(WORKBOOK).draft.active_plan

    assert active_plan.milestones["正式立项"].start_date == date(2026, 7, 30)
    assert active_plan.milestones["正式立项"].end_date == date(2026, 7, 30)
    assert active_plan.milestones["EVT可靠性测试"].start_date == date(2026, 9, 11)
    assert active_plan.milestones["EVT可靠性测试"].end_date == date(2026, 9, 26)


def test_xlsx_05_resolves_raci_formula_references_without_cached_values(
    registry: ParserRegistry,
) -> None:
    milestone = registry.parse(WORKBOOK).draft.milestones[1]

    assert milestone.name == "MD设计"
    assert milestone.assignments[RaciRole.RESPONSIBLE] == ["成员03"]
    assert milestone.assignments[RaciRole.ACCOUNTABLE] == ["成员02"]
    assert milestone.assignments[RaciRole.CONSULTED] == ["成员05", "成员06"]


def test_xlsx_05_all_milestones_have_outputs_and_complete_raci(
    registry: ParserRegistry,
) -> None:
    milestones = registry.parse(WORKBOOK).draft.milestones

    assert all(milestone.output for milestone in milestones)
    assert all(set(milestone.assignments) == set(RaciRole) for milestone in milestones)
    assert all(milestone.assignments[role] for milestone in milestones for role in RaciRole)


@pytest.mark.parametrize(
    ("mutation", "expected_location"),
    [
        ("remove_sheet", "项目团队构成"),
        ("remove_header", "项目进度表!A2"),
    ],
)
def test_xlsx_06_reports_precise_missing_structure(
    registry: ParserRegistry,
    phase1_tmp_path: Path,
    mutation: str,
    expected_location: str,
) -> None:
    changed = phase1_tmp_path / "invalid.xlsx"
    shutil.copyfile(WORKBOOK, changed)
    workbook = load_workbook(changed)
    if mutation == "remove_sheet":
        del workbook["项目团队构成"]
    else:
        workbook["项目进度表"]["A2"] = ""
    workbook.save(changed)

    with pytest.raises(WorkbookValidationError, match=expected_location):
        registry.parse(changed)


def test_xlsx_allows_missing_optional_customer_product_specification_sheet(
    registry: ParserRegistry,
    phase1_tmp_path: Path,
) -> None:
    changed = phase1_tmp_path / "without-customer-product-specification.xlsx"
    shutil.copyfile(WORKBOOK, changed)
    workbook = load_workbook(changed)
    del workbook["客户-产品规格书"]
    workbook.save(changed)
    workbook.close()

    result = registry.parse(changed)

    assert result.draft.project.code == EXPECTED["project"]["code"]
    assert len(result.draft.product_specs) == 70
    assert len(result.draft.milestones) == 24


@pytest.mark.parametrize("filename", ["legacy.xls", "disguised.xlsx"])
def test_xlsx_07_rejects_xls_and_disguised_files(
    registry: ParserRegistry,
    phase1_tmp_path: Path,
    filename: str,
) -> None:
    invalid = phase1_tmp_path / filename
    invalid.write_bytes(b"not an Office Open XML workbook")

    with pytest.raises(InvalidWorkbookError):
        registry.parse(invalid)


def test_xlsx_rejects_archives_that_expand_beyond_the_configured_limit(
    registry: ParserRegistry,
    phase1_tmp_path: Path,
) -> None:
    oversized = phase1_tmp_path / "oversized.xlsx"
    with zipfile.ZipFile(oversized, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("xl/worksheets/sheet1.xml", "x" * 2048)

    with pytest.raises(InvalidWorkbookError, match="expanded size limit"):
        registry.parse(
            oversized,
            max_uncompressed_size_bytes=1024,
            max_archive_entries=10,
        )


def test_xlsx_isolated_parser_enforces_execution_timeout(registry: ParserRegistry) -> None:
    with pytest.raises(InvalidWorkbookError, match="timed out"):
        registry.parse_isolated(
            WORKBOOK,
            timeout_seconds=0.001,
            max_uncompressed_size_bytes=50 * 1024 * 1024,
            max_archive_entries=1000,
        )


def test_xlsx_08_rejects_unknown_template_and_lists_supported_versions(
    registry: ParserRegistry,
    phase1_tmp_path: Path,
) -> None:
    unknown = phase1_tmp_path / "unknown.xlsx"
    shutil.copyfile(WORKBOOK, unknown)
    workbook = load_workbook(unknown)
    workbook["产品规格书"]["A1"] = "未知供应商模板"
    workbook["产品规格书"]["D4"] = "其他项目"
    workbook.save(unknown)

    with pytest.raises(TemplateNotSupportedError) as exc_info:
        registry.parse(unknown)

    assert exc_info.value.supported_versions == ["lyra_project_spec/1.0"]


def test_same_template_accepts_a_different_project_identity(
    registry: ParserRegistry,
    phase1_tmp_path: Path,
) -> None:
    changed = phase1_tmp_path / "different-project.xlsx"
    shutil.copyfile(WORKBOOK, changed)
    workbook = load_workbook(changed)
    workbook["产品规格书"]["D4"] = "Nova Pro（产品代号：Nova；项目号：ZPD9999；G300版本）"
    workbook.save(changed)

    draft = registry.parse(changed).draft

    assert draft.project.code == "ZPD9999"
    assert draft.project.name == "Nova Pro"


def test_same_person_with_multiple_team_roles_is_merged(
    registry: ParserRegistry,
    phase1_tmp_path: Path,
) -> None:
    changed = phase1_tmp_path / "multiple-roles.xlsx"
    shutil.copyfile(WORKBOOK, changed)
    workbook = load_workbook(changed)
    team = workbook["项目团队构成"]
    team["C6"] = team["C5"].value
    team["D6"] = team["D5"].value
    workbook.save(changed)

    draft = registry.parse(changed).draft

    member = next(item for item in draft.members if item.name == team["C5"].value)
    assert member.role == f"{team['B5'].value} / {team['B6'].value}"
    assert len(draft.members) == 21


def test_same_person_with_conflicting_phone_numbers_is_rejected(
    registry: ParserRegistry,
    phase1_tmp_path: Path,
) -> None:
    changed = phase1_tmp_path / "conflicting-member-phone.xlsx"
    shutil.copyfile(WORKBOOK, changed)
    workbook = load_workbook(changed)
    team = workbook["项目团队构成"]
    team["C6"] = team["C5"].value
    team["D5"] = "13800008888"
    team["D6"] = "13900009999"
    workbook.save(changed)

    with pytest.raises(WorkbookValidationError, match="conflicting contact details"):
        registry.parse(changed)


@pytest.mark.parametrize(
    ("manager_roles", "expected_message"),
    [
        (("结构经理",), "exactly one project manager is required"),
        (("项目经理", "项目经理"), "exactly one project manager is required"),
    ],
)
def test_team_requires_exactly_one_project_manager(
    registry: ParserRegistry,
    phase1_tmp_path: Path,
    manager_roles: tuple[str, ...],
    expected_message: str,
) -> None:
    changed = phase1_tmp_path / "invalid-project-manager.xlsx"
    shutil.copyfile(WORKBOOK, changed)
    workbook = load_workbook(changed)
    team = workbook["项目团队构成"]
    team["B5"] = manager_roles[0]
    if len(manager_roles) == 2:
        team["B6"] = manager_roles[1]
    workbook.save(changed)

    with pytest.raises(WorkbookValidationError, match=expected_message):
        registry.parse(changed)


def test_semantic_validation_rejects_unknown_raci_member(
    registry: ParserRegistry,
    phase1_tmp_path: Path,
) -> None:
    changed = phase1_tmp_path / "unknown-raci-member.xlsx"
    shutil.copyfile(WORKBOOK, changed)
    workbook = load_workbook(changed)
    workbook["项目进度表"]["B11"] = "不存在成员"
    workbook.save(changed)

    with pytest.raises(WorkbookValidationError, match="项目进度表!B11"):
        registry.parse(changed)


def _serialize_plan_window(window: PlanWindow) -> str | None:
    if window.state is PlanDateState.TBD:
        return None
    if window.state is PlanDateState.NOT_APPLICABLE:
        return "N/A"
    start_date = window.start_date
    end_date = window.end_date
    assert start_date is not None and end_date is not None
    if start_date == end_date:
        return start_date.isoformat()
    return f"{start_date.isoformat()}/{end_date.isoformat()}"
